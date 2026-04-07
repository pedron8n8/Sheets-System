"""Motor de formulas do workbook.

Visao geral do que este arquivo cobre:
- Aba Inputs: recalculo de formulas derivadas (Sources & Uses, Refi, Exit, Waterfall params).
- Aba IncomeExpenses: leitura de labels/valores para alimentar linhas dinamicas no Monthly CF.
- Aba Monthly CF: calendario (mes/trimestre/ano), formulas mensais, divida e distribuicoes.
- Abas Quarterly CF e Annual CF: agregacao do Monthly CF por periodo.
- Aba Summary: apontamento de Exit Cap Rate.
- Aba Equity Waterfall: reconstrucao de formulas sem depender de linhas fixas.

Por que existe:
- O template muda de versao e linhas podem deslocar.
- Para evitar quebrar formulas, o codigo procura labels e recalcula referencias dinamicamente.
"""

from __future__ import annotations

import argparse
import copy
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PASTA_OUTPUT = Path("Output")
ARQUIVO_TEMPLATE = Path("InputTemplate.xlsx")
ABA_PADRAO = "Monthly CF"
ABA_INCOME_EXPENSES = "IncomeExpenses"
IE_INCOME_START = 7
IE_INCOME_END = 46
IE_EXPENSE_START = 50
IE_EXPENSE_END = 89
BASE_OTHER_INCOME_SLOTS = 5
BASE_EXPENSE_SLOTS = 15


def _set_cell_value_respeitando_merge(sheet, cell_ref: str, value) -> None:
    """Escreve valor mesmo quando a celula alvo esta em merge.

Onde atua:
- Qualquer aba do workbook (Inputs, Monthly CF, Summary etc.).

Por que faz:
- O openpyxl so permite escrita segura na ancora (canto superior esquerdo) do merge.
    """
    for merged_range in sheet.merged_cells.ranges:
        if cell_ref in merged_range:
            anchor_col = get_column_letter(merged_range.min_col)
            anchor_ref = f"{anchor_col}{merged_range.min_row}"
            sheet[anchor_ref] = value
            return
    sheet[cell_ref] = value


def _normalizar_purchase_date(valor):
    """Normaliza Purchase Date para objeto date.

Onde atua:
- Campo de data vindo da aba Inputs (usado para montar calendario da Monthly CF).

Por que faz:
- A data pode vir como datetime, date ou string em formatos diferentes.
    """
    if valor is None or valor == "":
        return None

    if isinstance(valor, datetime):
        return valor.date()

    if isinstance(valor, date):
        return valor

    if isinstance(valor, str):
        texto = valor.strip()
        if not texto:
            return None
        for formato in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(texto, formato).date()
            except ValueError:
                continue

    return None


def _normalizar_end_year(valor, padrao: int = 10) -> int:
    """Normaliza End Year para inteiro positivo.

Onde atua:
- Campo End Year da aba Inputs.

Por que faz:
- Evita horizonte invalido (nulo, texto ou valor <= 0) quebrando formulas de timeline.
    """
    if valor is None or valor == "":
        return padrao

    try:
        ano = int(float(valor))
    except (TypeError, ValueError):
        return padrao

    return ano if ano > 0 else padrao


def _texto_normalizado(valor) -> str:
    """Normaliza texto para comparacoes de labels.

Onde atua:
- Buscas por labels nas abas Inputs, Monthly CF, Quarterly CF e Annual CF.

Por que faz:
- Remove variacao de espacos/maiuculas e NBSP para match mais robusto.
    """
    if not isinstance(valor, str):
        return ""
    return " ".join(valor.replace("\xa0", " ").strip().lower().split())


def _encontrar_linha_por_texto(ws, texto: str, linha_inicial: int = 1) -> int | None:
    """Encontra a primeira linha cujo texto da coluna B contenha o alvo.

Onde atua:
- Principalmente Monthly CF, Quarterly CF e Annual CF para localizar blocos/totalizadores.

Por que faz:
- Permite ancorar formulas por label em vez de numero fixo de linha.
    """
    alvo = texto.lower()
    for linha in range(max(1, linha_inicial), ws.max_row + 1):
        valor_b = _texto_normalizado(ws.cell(row=linha, column=2).value)
        if alvo in valor_b:
            return linha
    return None


def _encontrar_linha_input_por_label(
    ws_inputs,
    label: str,
    linha_inicial: int = 1,
    linha_final: int | None = None,
    correspondencia_exata: bool = True,
    ultima_ocorrencia: bool = False,
) -> int | None:
    """Busca linha na aba Inputs usando label na coluna B.

Onde atua:
- Aba Inputs.

Por que faz:
- Inputs pode ter labels duplicados em secoes diferentes; os parametros permitem
    escolher escopo, match exato/parcial e primeira/ultima ocorrencia.
        """
    alvo = _texto_normalizado(label)
    if not alvo:
        return None

    inicio = max(1, linha_inicial)
    fim = ws_inputs.max_row if linha_final is None else min(ws_inputs.max_row, max(1, linha_final))

    encontrada = None
    for linha in range(inicio, fim + 1):
        texto_linha = _texto_normalizado(ws_inputs.cell(row=linha, column=2).value)
        if not texto_linha:
            continue

        bate = texto_linha == alvo if correspondencia_exata else alvo in texto_linha
        if not bate:
            continue

        if not ultima_ocorrencia:
            return linha
        encontrada = linha

    return encontrada


def _obter_linha_input_obrigatoria(
    ws_inputs,
    label: str,
    linha_inicial: int = 1,
    linha_final: int | None = None,
    correspondencia_exata: bool = True,
    ultima_ocorrencia: bool = False,
) -> int:
    """Versao obrigatoria de busca de label na Inputs.

Onde atua:
- Aba Inputs.

Por que faz:
- Falha cedo com erro descritivo quando um label critico nao existe no template.
    """
    linha = _encontrar_linha_input_por_label(
        ws_inputs,
        label,
        linha_inicial=linha_inicial,
        linha_final=linha_final,
        correspondencia_exata=correspondencia_exata,
        ultima_ocorrencia=ultima_ocorrencia,
    )
    if linha is None:
        raise ValueError(f"Label nao encontrado no Inputs: {label}")
    return linha


def _copiar_estilo_linha(ws, linha_origem: int, linha_destino: int) -> None:
    """Replica estilo visual de uma linha para outra.

Onde atua:
- Monthly CF, Quarterly CF e Annual CF quando linhas dinamicas sao inseridas.

Por que faz:
- Novas linhas precisam manter formato do template (borda, fonte, numero etc.).
    """
    for col in range(1, ws.max_column + 1):
        origem = ws.cell(row=linha_origem, column=col)
        destino = ws.cell(row=linha_destino, column=col)
        destino._style = copy.copy(origem._style)


def _capturar_merges(ws) -> list[tuple[int, int, int, int]]:
    """Captura todos os intervalos mesclados de uma aba."""
    return [(m.min_row, m.max_row, m.min_col, m.max_col) for m in ws.merged_cells.ranges]


def _remover_todos_merges(ws) -> None:
    """Remove temporariamente todos os merges da aba.

Por que faz:
- Insercao de linhas com merges ativos costuma deslocar referencias de forma incorreta.
    """
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))


def _deslocar_intervalo_por_insercoes(
    min_row: int,
    max_row: int,
    insercoes: list[tuple[int, int]],
) -> tuple[int, int]:
    """Recalcula inicio/fim de um merge apos insercoes de linhas."""
    novo_min = min_row
    novo_max = max_row
    for linha_insercao, quantidade in sorted(insercoes, key=lambda x: x[0]):
        if linha_insercao <= novo_min:
            novo_min += quantidade
            novo_max += quantidade
        elif novo_min < linha_insercao <= novo_max:
            novo_max += quantidade
    return novo_min, novo_max


def _reaplicar_merges_deslocados(
    ws,
    merges_originais: list[tuple[int, int, int, int]],
    insercoes: list[tuple[int, int]],
) -> None:
    """Aplica novamente merges com coordenadas ja deslocadas."""
    for min_row, max_row, min_col, max_col in merges_originais:
        novo_min, novo_max = _deslocar_intervalo_por_insercoes(min_row, max_row, insercoes)
        ws.merge_cells(
            start_row=novo_min,
            end_row=novo_max,
            start_column=min_col,
            end_column=max_col,
        )


def _insert_rows_preservando_merges(ws, linha_insert: int, quantidade: int) -> None:
    """Insere linhas preservando merges existentes.

Onde atua:
- Blocos dinamicos de Monthly CF, Quarterly CF e Annual CF.

Por que faz:
- Evita corrupcao visual/estrutural quando o layout cresce para suportar novos itens.
    """
    if quantidade <= 0:
        return

    merges_originais = _capturar_merges(ws)
    _remover_todos_merges(ws)
    ws.insert_rows(linha_insert, quantidade)
    _reaplicar_merges_deslocados(ws, merges_originais, [(linha_insert, quantidade)])


def _obter_layout_monthly_cf(ws_monthly) -> dict:
    """Mapeia linhas-chave da aba Monthly CF por label.

Onde atua:
- Aba Monthly CF.

O que faz:
- Localiza headers, subtotais e blocos (receita, despesas, aquisicao, dividas,
  distribuicoes e metricas acumuladas).
- Retorna um dicionario com todas as linhas necessarias para escrever formulas.

Por que faz:
- O layout pode mudar entre templates; mapear por texto evita dependencia de linhas fixas.
    """
    row_total_revenue = _encontrar_linha_por_texto(ws_monthly, "total revenue", 1)
    row_op_exp_header = _encontrar_linha_por_texto(ws_monthly, "operating expenses", row_total_revenue or 1)
    row_total_expenses = _encontrar_linha_por_texto(ws_monthly, "total operating expenses", row_op_exp_header or 1)
    row_noi = _encontrar_linha_por_texto(ws_monthly, "net operating income", row_total_expenses or 1)
    row_capex = _encontrar_linha_por_texto(ws_monthly, "capital expenditures (capex)", row_noi or 1)
    row_total_acquisition = _encontrar_linha_por_texto(ws_monthly, "total acquisition costs", row_capex or 1)
    row_sale_proceeds = _encontrar_linha_por_texto(ws_monthly, "sale proceeds", row_total_acquisition or 1)
    row_selling_costs = _encontrar_linha_por_texto(ws_monthly, "selling costs", row_sale_proceeds or 1)
    row_net_sale = _encontrar_linha_por_texto(ws_monthly, "net sale proceeds", row_selling_costs or 1)
    row_cf_before_debt = _encontrar_linha_por_texto(ws_monthly, "cashflow before debt service", row_net_sale or 1)
    row_sf_header = _encontrar_linha_por_texto(ws_monthly, "seller finance", row_cf_before_debt or 1)
    row_bl_header = _encontrar_linha_por_texto(ws_monthly, "bank loan", row_sf_header or 1)
    row_rf_header = _encontrar_linha_por_texto(ws_monthly, "refinance loan", row_bl_header or 1)
    row_equity = _encontrar_linha_por_texto(ws_monthly, "equity contribution (inflow)", row_rf_header or 1)
    row_refi_net = _encontrar_linha_por_texto(ws_monthly, "refi proceeds (net)", row_equity or 1)
    row_loan_proceeds = _encontrar_linha_por_texto(ws_monthly, "loan proceeds (inflow)", row_refi_net or 1)
    row_total_debt = _encontrar_linha_por_texto(ws_monthly, "total debt service", row_loan_proceeds or 1)
    row_cf_after_debt = _encontrar_linha_por_texto(ws_monthly, "cashflow after debt service", row_total_debt or 1)
    row_cum_cf = _encontrar_linha_por_texto(ws_monthly, "cumulative cashflow", row_cf_after_debt or 1)
    row_member_header = _encontrar_linha_por_texto(ws_monthly, "member distributions", row_cum_cf or 1)
    row_asset_fee = _encontrar_linha_por_texto(ws_monthly, "asset management fee", row_member_header or 1)
    row_cash_members = _encontrar_linha_por_texto(ws_monthly, "cash available for members", row_asset_fee or 1)
    row_total_dist = _encontrar_linha_por_texto(ws_monthly, "total distributions", row_cash_members or 1)
    row_cum_metrics = _encontrar_linha_por_texto(ws_monthly, "cumulative metrics", row_total_dist or 1)
    row_cum_noi = _encontrar_linha_por_texto(ws_monthly, "cumulative noi", row_cum_metrics or 1)
    row_cum_cf_after_debt = _encontrar_linha_por_texto(ws_monthly, "cumulative cf after debt", row_cum_noi or 1)
    row_dscr = _encontrar_linha_por_texto(ws_monthly, "dscr", row_cum_cf_after_debt or 1)

    obrigatorias = [
        row_total_revenue,
        row_op_exp_header,
        row_total_expenses,
        row_noi,
        row_capex,
        row_total_acquisition,
        row_sale_proceeds,
        row_selling_costs,
        row_net_sale,
        row_cf_before_debt,
        row_sf_header,
        row_bl_header,
        row_rf_header,
        row_equity,
        row_refi_net,
        row_loan_proceeds,
        row_total_debt,
        row_cf_after_debt,
        row_cum_cf,
        row_asset_fee,
        row_cash_members,
        row_total_dist,
        row_cum_metrics,
        row_cum_noi,
        row_cum_cf_after_debt,
        row_dscr,
    ]
    if any(l is None for l in obrigatorias):
        raise ValueError("Layout da aba Monthly CF nao encontrado para formulas dinamicas.")

    return {
        "income_start": 12,
        "income_end": row_total_revenue - 1,
        "row_total_revenue": row_total_revenue,
        "row_op_exp_header": row_op_exp_header,
        "expense_start": row_op_exp_header + 1,
        "expense_end": row_total_expenses - 1,
        "row_total_expenses": row_total_expenses,
        "row_noi": row_noi,
        "row_capex": row_capex,
        "row_purchase": row_total_acquisition - 6,
        "row_closing": row_total_acquisition - 5,
        "row_immediate": row_total_acquisition - 4,
        "row_acq_fee": row_total_acquisition - 3,
        "row_due_diligence": row_total_acquisition - 2,
        "row_loan_origination": row_total_acquisition - 1,
        "row_total_acquisition": row_total_acquisition,
        "row_sale_proceeds": row_sale_proceeds,
        "row_selling_costs": row_selling_costs,
        "row_net_sale": row_net_sale,
        "row_cf_before_debt": row_cf_before_debt,
        "row_sf_header": row_sf_header,
        "row_sf_begin": row_sf_header + 1,
        "row_sf_funding": row_sf_header + 2,
        "row_sf_interest": row_sf_header + 3,
        "row_sf_principal": row_sf_header + 4,
        "row_sf_payoff": row_sf_header + 5,
        "row_sf_end": row_sf_header + 6,
        "row_sf_total": row_sf_header + 7,
        "row_bl_header": row_bl_header,
        "row_bl_begin": row_bl_header + 1,
        "row_bl_funding": row_bl_header + 2,
        "row_bl_interest": row_bl_header + 3,
        "row_bl_principal": row_bl_header + 4,
        "row_bl_payoff": row_bl_header + 5,
        "row_bl_end": row_bl_header + 6,
        "row_bl_total": row_bl_header + 7,
        "row_rf_header": row_rf_header,
        "row_rf_begin": row_rf_header + 1,
        "row_rf_funding": row_rf_header + 2,
        "row_rf_interest": row_rf_header + 3,
        "row_rf_principal": row_rf_header + 4,
        "row_rf_payoff": row_rf_header + 5,
        "row_rf_end": row_rf_header + 6,
        "row_rf_total": row_rf_header + 7,
        "row_equity": row_equity,
        "row_refi_net": row_refi_net,
        "row_loan_proceeds": row_loan_proceeds,
        "row_total_debt": row_total_debt,
        "row_cf_after_debt": row_cf_after_debt,
        "row_cum_cf": row_cum_cf,
        "row_asset_fee": row_asset_fee,
        "row_cash_members": row_cash_members,
        "row_cap_partner_dist": row_cash_members + 1,
        "row_return_capital": row_cash_members + 2,
        "row_manager_dist": row_cash_members + 3,
        "row_total_dist": row_total_dist,
        "row_cum_noi": row_cum_noi,
        "row_cum_cf_after_debt": row_cum_cf_after_debt,
        "row_dscr": row_dscr,
    }


def _obter_layout_resumo_cf(ws_resumo) -> dict:
    """Mapeia linhas-chave das abas Quarterly CF e Annual CF.

Onde atua:
- Aba Quarterly CF ou Annual CF.

O que faz:
- Detecta estrutura equivalente ao Monthly em visao agregada e retorna linhas
  de receita, despesa, divida e distribuicao.

Por que faz:
- Permite que as formulas de resumo sejam reaplicadas mesmo com deslocamentos.
    """
    row_revenue_header = _encontrar_linha_por_texto(ws_resumo, "revenue", 1)
    row_total_revenue = _encontrar_linha_por_texto(ws_resumo, "total revenue", row_revenue_header or 1)
    row_op_exp_header = _encontrar_linha_por_texto(ws_resumo, "operating expenses", row_total_revenue or 1)
    row_total_expenses = _encontrar_linha_por_texto(ws_resumo, "total operating expenses", row_op_exp_header or 1)
    row_noi = _encontrar_linha_por_texto(ws_resumo, "net operating income", row_total_expenses or 1)
    row_capex = _encontrar_linha_por_texto(ws_resumo, "capital expenditures", row_noi or 1)
    row_acquisition = _encontrar_linha_por_texto(ws_resumo, "acquisition costs", row_capex or 1)
    row_sale = _encontrar_linha_por_texto(ws_resumo, "sale proceeds", row_acquisition or 1)
    row_selling = _encontrar_linha_por_texto(ws_resumo, "selling costs", row_sale or 1)
    row_net_sale = _encontrar_linha_por_texto(ws_resumo, "net sale proceeds", row_selling or 1)
    row_cf_before_debt = _encontrar_linha_por_texto(ws_resumo, "cashflow before debt service", row_net_sale or 1)
    row_sf_header = _encontrar_linha_por_texto(ws_resumo, "seller finance", row_cf_before_debt or 1)
    row_bl_header = _encontrar_linha_por_texto(ws_resumo, "bank loan", row_sf_header or 1)
    row_rf_header = _encontrar_linha_por_texto(ws_resumo, "refinance loan", row_bl_header or 1)
    row_equity = _encontrar_linha_por_texto(ws_resumo, "equity contribution", row_rf_header or 1)
    row_refi_net = _encontrar_linha_por_texto(ws_resumo, "refi proceeds (net)", row_equity or 1)
    row_loan_proceeds = _encontrar_linha_por_texto(ws_resumo, "loan proceeds", row_refi_net or 1)
    row_total_debt = _encontrar_linha_por_texto(ws_resumo, "total debt service", row_loan_proceeds or 1)
    row_cf_after_debt = _encontrar_linha_por_texto(ws_resumo, "cashflow after debt service", row_total_debt or 1)
    row_cum_cf = _encontrar_linha_por_texto(ws_resumo, "cumulative cashflow", row_cf_after_debt or 1)
    row_asset_fee = _encontrar_linha_por_texto(ws_resumo, "asset management fee", row_cum_cf or 1)
    if row_asset_fee is None:
        row_asset_fee = _encontrar_linha_por_texto(ws_resumo, "asset mgmt fee", row_cum_cf or 1)
    row_cash_members = _encontrar_linha_por_texto(ws_resumo, "cash available for members", row_asset_fee or 1)
    row_total_dist = _encontrar_linha_por_texto(ws_resumo, "total distributions", row_cash_members or 1)
    row_dscr = _encontrar_linha_por_texto(ws_resumo, "dscr", row_total_dist or 1)

    obrigatorias = [
        row_revenue_header,
        row_total_revenue,
        row_op_exp_header,
        row_total_expenses,
        row_noi,
        row_capex,
        row_acquisition,
        row_sale,
        row_selling,
        row_net_sale,
        row_cf_before_debt,
        row_sf_header,
        row_bl_header,
        row_rf_header,
        row_equity,
        row_refi_net,
        row_loan_proceeds,
        row_total_debt,
        row_cf_after_debt,
        row_cum_cf,
        row_asset_fee,
        row_cash_members,
        row_total_dist,
        row_dscr,
    ]
    if any(l is None for l in obrigatorias):
        raise ValueError("Layout da aba de resumo CF nao encontrado para formulas dinamicas.")

    return {
        "row_revenue_header": row_revenue_header,
        "row_gross": row_revenue_header + 1,
        "row_vacancy": row_revenue_header + 2,
        "row_credit_loss": row_revenue_header + 3,
        "row_egi": row_revenue_header + 4,
        "income_start": row_revenue_header + 5,
        "income_end": row_total_revenue - 1,
        "row_total_revenue": row_total_revenue,
        "row_op_exp_header": row_op_exp_header,
        "expense_start": row_op_exp_header + 1,
        "expense_end": row_total_expenses - 1,
        "row_total_expenses": row_total_expenses,
        "row_noi": row_noi,
        "row_capex": row_capex,
        "row_acquisition": row_acquisition,
        "row_sale": row_sale,
        "row_selling": row_selling,
        "row_net_sale": row_net_sale,
        "row_cf_before_debt": row_cf_before_debt,
        "row_sf_header": row_sf_header,
        "row_sf_begin": row_sf_header + 1,
        "row_sf_funding": row_sf_header + 2,
        "row_sf_interest": row_sf_header + 3,
        "row_sf_principal": row_sf_header + 4,
        "row_sf_payoff": row_sf_header + 5,
        "row_sf_end": row_sf_header + 6,
        "row_sf_total": row_sf_header + 7,
        "row_bl_header": row_bl_header,
        "row_bl_begin": row_bl_header + 1,
        "row_bl_funding": row_bl_header + 2,
        "row_bl_interest": row_bl_header + 3,
        "row_bl_principal": row_bl_header + 4,
        "row_bl_payoff": row_bl_header + 5,
        "row_bl_end": row_bl_header + 6,
        "row_bl_total": row_bl_header + 7,
        "row_rf_header": row_rf_header,
        "row_rf_begin": row_rf_header + 1,
        "row_rf_funding": row_rf_header + 2,
        "row_rf_interest": row_rf_header + 3,
        "row_rf_principal": row_rf_header + 4,
        "row_rf_payoff": row_rf_header + 5,
        "row_rf_end": row_rf_header + 6,
        "row_rf_total": row_rf_header + 7,
        "row_equity": row_equity,
        "row_refi_net": row_refi_net,
        "row_loan_proceeds": row_loan_proceeds,
        "row_total_debt": row_total_debt,
        "row_cf_after_debt": row_cf_after_debt,
        "row_cum_cf": row_cum_cf,
        "row_asset_fee": row_asset_fee,
        "row_cash_members": row_cash_members,
        "row_cap_partner_dist": row_cash_members + 1,
        "row_return_capital": row_cash_members + 2,
        "row_manager_dist": row_cash_members + 3,
        "row_total_dist": row_total_dist,
        "row_dscr": row_dscr,
    }


def _expandir_linhas_monthly_cf(ws_monthly, minimo_income_slots: int, minimo_expense_slots: int) -> dict:
    """Retorna layout da Monthly CF sem expansao de linhas.

O template atual trabalha com blocos fixos para receitas e despesas,
entao esta funcao e intencionalmente no-op para preservar a estrutura.
    """
    _ = (minimo_income_slots, minimo_expense_slots)
    return _obter_layout_monthly_cf(ws_monthly)


def _expandir_linhas_resumo_cf(ws_resumo, minimo_income_slots: int, minimo_expense_slots: int) -> dict:
    """Retorna layout de Quarterly/Annual CF sem expansao de linhas.

Mantemos comportamento no-op para acompanhar o template com slots fixos.
    """
    _ = (minimo_income_slots, minimo_expense_slots)
    return _obter_layout_resumo_cf(ws_resumo)


def _extrair_nomes_other_income(ws_ie) -> list[str]:
    """Le nomes de Other Income na aba IncomeExpenses (coluna B, linhas 7-46)."""
    nomes = []
    for linha in range(IE_INCOME_START, IE_INCOME_END + 1):
        valor_b = ws_ie.cell(row=linha, column=2).value
        if isinstance(valor_b, str):
            nome = valor_b.strip()
            if nome:
                nomes.append(nome)
    return nomes


def _extrair_nomes_expenses(ws_ie) -> list[str]:
    """Le nomes de Expense na aba IncomeExpenses (coluna B, linhas 50-89)."""
    nomes = []
    for linha in range(IE_EXPENSE_START, IE_EXPENSE_END + 1):
        valor_b = ws_ie.cell(row=linha, column=2).value
        if isinstance(valor_b, str):
            nome = valor_b.strip()
            if nome:
                nomes.append(nome)
    return nomes


def _atualizar_nomes_other_income_monthly_cf(
    ws_monthly,
    nomes: list[str],
    linha_inicial: int = 12,
    linha_final: int | None = None,
) -> None:
    """Copia labels de Other Income para a area dinamica da Monthly CF.

Onde atua:
- Origem: IncomeExpenses.
- Destino: Monthly CF (coluna B, bloco de receitas adicionais).

Por que faz:
- As formulas no Monthly usam o label para buscar valor correspondente na IncomeExpenses.
    """
    if linha_final is None:
        layout = _obter_layout_monthly_cf(ws_monthly)
        linha_final = layout["income_end"]

    cursor = 0
    for linha in range(linha_inicial, linha_final + 1):
        if cursor < len(nomes):
            ws_monthly.cell(row=linha, column=2, value=nomes[cursor])
            cursor += 1
        else:
            ws_monthly.cell(row=linha, column=2, value="")


def _atualizar_nomes_expenses_monthly_cf(
    ws_monthly,
    nomes: list[str],
    linha_inicial: int | None = None,
    linha_final: int | None = None,
) -> None:
    """Copia labels de despesas para a area dinamica da Monthly CF.

Onde atua:
- Origem: IncomeExpenses.
- Destino: Monthly CF (coluna B, bloco de despesas operacionais).
    """
    layout = _obter_layout_monthly_cf(ws_monthly)
    if linha_inicial is None:
        linha_inicial = layout["expense_start"]
    if linha_final is None:
        linha_final = layout["expense_end"]

    cursor = 0
    for linha in range(linha_inicial, linha_final + 1):
        if cursor < len(nomes):
            ws_monthly.cell(row=linha, column=2, value=nomes[cursor])
            cursor += 1
        else:
            ws_monthly.cell(row=linha, column=2, value="")


def _aplicar_formulas_monthly_cf_dinamicas(ws_monthly, ws_inputs, ws_ie) -> None:
    """Reconstroi todas as formulas da aba Monthly CF de forma dinamica.

Onde atua:
- Aba Monthly CF (destino das formulas).
- Aba Inputs (parametros: prazos, taxas, valores, shares).
- Aba IncomeExpenses (valores base de receitas/despesas por label).

O que faz:
- Preenche formulas de receita/despesa mensal.
- Monta bloco de aquisicao, venda e cashflow.
- Monta blocos de Seller Finance, Bank Loan e Refi Loan.
- Calcula distribuicoes e metricas (cumulativos, DSCR).

Por que faz:
- Evita formulas quebradas por mudanca de linha no template,
  sempre ancorando por labels mapeados dinamicamente.
    """
    layout = _obter_layout_monthly_cf(ws_monthly)

    row_income_start = layout["income_start"]
    row_income_end = layout["income_end"]
    row_total_revenue = layout["row_total_revenue"]
    row_egi = 11

    row_expense_start = layout["expense_start"]
    row_expense_end = layout["expense_end"]
    row_total_expenses = layout["row_total_expenses"]
    row_noi = layout["row_noi"]
    row_capex = layout["row_capex"]

    row_purchase = layout["row_purchase"]
    row_closing = layout["row_closing"]
    row_immediate = layout["row_immediate"]
    row_acq_fee = layout["row_acq_fee"]
    row_due_diligence = layout["row_due_diligence"]
    row_loan_origination = layout["row_loan_origination"]
    row_total_acquisition = layout["row_total_acquisition"]

    row_sale_proceeds = layout["row_sale_proceeds"]
    row_selling_costs = layout["row_selling_costs"]
    row_net_sale = layout["row_net_sale"]
    row_cf_before_debt = layout["row_cf_before_debt"]

    row_sf_begin = layout["row_sf_begin"]
    row_sf_funding = layout["row_sf_funding"]
    row_sf_interest = layout["row_sf_interest"]
    row_sf_principal = layout["row_sf_principal"]
    row_sf_payoff = layout["row_sf_payoff"]
    row_sf_end = layout["row_sf_end"]
    row_sf_total = layout["row_sf_total"]

    row_bl_begin = layout["row_bl_begin"]
    row_bl_funding = layout["row_bl_funding"]
    row_bl_interest = layout["row_bl_interest"]
    row_bl_principal = layout["row_bl_principal"]
    row_bl_payoff = layout["row_bl_payoff"]
    row_bl_end = layout["row_bl_end"]
    row_bl_total = layout["row_bl_total"]

    row_rf_begin = layout["row_rf_begin"]
    row_rf_funding = layout["row_rf_funding"]
    row_rf_interest = layout["row_rf_interest"]
    row_rf_principal = layout["row_rf_principal"]
    row_rf_payoff = layout["row_rf_payoff"]
    row_rf_end = layout["row_rf_end"]
    row_rf_total = layout["row_rf_total"]

    row_equity = layout["row_equity"]
    row_refi_net = layout["row_refi_net"]
    row_loan_proceeds = layout["row_loan_proceeds"]
    row_total_debt = layout["row_total_debt"]
    row_cf_after_debt = layout["row_cf_after_debt"]
    row_cum_cf = layout["row_cum_cf"]
    row_asset_fee = layout["row_asset_fee"]
    row_cash_members = layout["row_cash_members"]
    row_cap_partner_dist = layout["row_cap_partner_dist"]
    row_return_capital = layout["row_return_capital"]
    row_manager_dist = layout["row_manager_dist"]
    row_total_dist = layout["row_total_dist"]
    row_cum_noi = layout["row_cum_noi"]
    row_cum_cf_after_debt = layout["row_cum_cf_after_debt"]
    row_dscr = layout["row_dscr"]
    # Linhas com comportamento fixo do template na Monthly CF.
    linhas_protegidas_template = {12, 16, 17}
    sheet_ie_ref = f"'{ws_ie.title}'" if ws_ie is not None else f"'{ABA_INCOME_EXPENSES}'"

    # Labels fixos exigidos pelo template nas linhas de resumo/agrupamento.
    ws_monthly.cell(row=12, column=2, value="Revenue")
    ws_monthly.cell(row=16, column=2, value="Property Management Fee (%EGI)")
    ws_monthly.cell(row=17, column=2, value="Expenses")

    # Linhas da Inputs sao resolvidas por label para evitar referencias fixas.
    row_input_closing_costs = _obter_linha_input_obrigatoria(
        ws_inputs,
        "Closing Costs",
        correspondencia_exata=True,
        ultima_ocorrencia=True,
    )
    row_input_immediate_repairs = _obter_linha_input_obrigatoria(
        ws_inputs,
        "Immediate Repairs / Reserves",
        correspondencia_exata=True,
        ultima_ocorrencia=True,
    )
    row_input_acquisition_fee_pct = _obter_linha_input_obrigatoria(
        ws_inputs,
        "Acquisition Fee (% Purchase)",
        correspondencia_exata=True,
        ultima_ocorrencia=True,
    )
    row_input_selling_price = _obter_linha_input_obrigatoria(
        ws_inputs,
        "Selling Price",
        correspondencia_exata=True,
        ultima_ocorrencia=True,
    )
    row_input_selling_costs = _obter_linha_input_obrigatoria(
        ws_inputs,
        "Selling Costs ($)",
        correspondencia_exata=True,
        ultima_ocorrencia=True,
    )
    row_input_equity_contribution = _obter_linha_input_obrigatoria(
        ws_inputs,
        "Equity Contribution",
        correspondencia_exata=True,
        ultima_ocorrencia=True,
    )
    row_input_asset_mgmt_fee = _obter_linha_input_obrigatoria(
        ws_inputs,
        "Asset Management Fee (% EGI)",
        correspondencia_exata=True,
        ultima_ocorrencia=True,
    )
    row_input_cap_partner_contribution = _obter_linha_input_obrigatoria(
        ws_inputs,
        "Capital Partner Contribution ($)",
        correspondencia_exata=True,
        ultima_ocorrencia=True,
    )
    row_input_cap_partner_share = _obter_linha_input_obrigatoria(
        ws_inputs,
        "Capital Partner Share (%)",
        correspondencia_exata=True,
        ultima_ocorrencia=True,
    )
    row_input_manager_share = _obter_linha_input_obrigatoria(
        ws_inputs,
        "Manager Share (%)",
        correspondencia_exata=True,
        ultima_ocorrencia=True,
    )

    ws_monthly.cell(row=row_due_diligence, column=2, value="Due Diligence Costs")
    ws_monthly.cell(row=row_loan_origination, column=2, value="Loan Origination Costs")

    for coluna in range(3, ws_monthly.max_column + 1):
        col = get_column_letter(coluna)
        col_anterior = get_column_letter(coluna - 1) if coluna > 3 else None

        # Receita bruta, vacancia e perdas por credito (linhas base do topo da Monthly CF).
        formula_gross = (
            f'=IF(AND({col}$3>=INDEX(Inputs!$C:$C,MATCH("*Revenue Start Month*",Inputs!$B:$B,0)),'
            f'{col}$3<=INDEX(Inputs!$C:$C,MATCH("End Year",Inputs!$B:$B,0))*12),'
            f'INDEX(Inputs!$E:$E,MATCH("*Gross Potential Rent*",Inputs!$B:$B,0))/12*'
            f'(1+INDEX(Inputs!$C:$C,MATCH("*Annual Revenue Growth Rate*",Inputs!$B:$B,0)))^({col}$5-1),0)'
        )
        formula_vacancy = (
            f'=IF({col}$3<=INDEX(Inputs!$C:$C,MATCH("End Year",Inputs!$B:$B,0))*12,'
            f'-{col}8*INDEX(Inputs!$E:$E,MATCH("*Vacancy*",Inputs!$B:$B,0)),0)'
        )
        formula_credit_loss = (
            f'=IF({col}$3<=INDEX(Inputs!$C:$C,MATCH("End Year",Inputs!$B:$B,0))*12,'
            f'-{col}8*INDEX(Inputs!$E:$E,MATCH("*Credit Loss*",Inputs!$B:$B,0)),0)'
        )

        ws_monthly.cell(row=8, column=coluna, value=formula_gross)
        ws_monthly.cell(row=9, column=coluna, value=formula_vacancy)
        ws_monthly.cell(row=10, column=coluna, value=formula_credit_loss)

        # Other Income dinamico: busca o valor anual em IncomeExpenses e converte para mensal.
        for linha_income in range(row_income_start, row_income_end + 1):
            formula_other_income = (
                f'=IF($B{linha_income}="",0,IF(AND({col}$3>=INDEX(Inputs!$C:$C,MATCH("*Revenue Start Month*",Inputs!$B:$B,0)),'
                f'{col}$3<=INDEX(Inputs!$C:$C,MATCH("End Year",Inputs!$B:$B,0))*12),'
                f'INDEX({sheet_ie_ref}!$E:$E,MATCH($B{linha_income},{sheet_ie_ref}!$B:$B,0))/12*'
                f'(1+INDEX(Inputs!$C:$C,MATCH("*Annual Revenue Growth Rate*",Inputs!$B:$B,0)))^({col}$5-1),0))'
            )
            if linha_income not in linhas_protegidas_template:
                ws_monthly.cell(row=linha_income, column=coluna, value=formula_other_income)

        formula_total_revenue = f'={col}{row_egi}+SUM({col}{row_income_start}:{col}{row_income_end})'
        if row_total_revenue not in linhas_protegidas_template:
            ws_monthly.cell(row=row_total_revenue, column=coluna, value=formula_total_revenue)

        # Expenses dinamicos: management fee usa EGI; demais despesas crescem por taxa anual.
        for linha_expense in range(row_expense_start, row_expense_end + 1):
            if linha_expense in linhas_protegidas_template:
                continue
            formula_expense = (
                f'=IF($B{linha_expense}="",0,IF(AND({col}$3>=INDEX(Inputs!$C:$C,MATCH("*Expense Start Month*",Inputs!$B:$B,0)),'
                f'{col}$3<=INDEX(Inputs!$C:$C,MATCH("End Year",Inputs!$B:$B,0))*12),'
                f'IF(ISNUMBER(SEARCH("management fee",$B{linha_expense})),'
                f'-{col}{row_egi}*INDEX({sheet_ie_ref}!$E:$E,MATCH($B{linha_expense},{sheet_ie_ref}!$B:$B,0)),'
                f'-INDEX({sheet_ie_ref}!$E:$E,MATCH($B{linha_expense},{sheet_ie_ref}!$B:$B,0))/12*'
                f'(1+INDEX(Inputs!$C:$C,MATCH("*Annual Expense Growth Rate*",Inputs!$B:$B,0)))^({col}$5-1)),0))'
            )
            ws_monthly.cell(row=linha_expense, column=coluna, value=formula_expense)

        # Formulas fixas do template: Revenue (linha 12), Property Management (16) e Expenses (17).
        formula_revenue_l12 = (
            f'=IF($B12="",0,IF(AND({col}$3>=INDEX(Inputs!$C:$C,MATCH("*Revenue Start Month*",Inputs!$B:$B,0)),'
            f'{col}$3<=INDEX(Inputs!$C:$C,MATCH("End Year",Inputs!$B:$B,0))*12),'
            f'IncomeExpenses!$C$47/12*(1+INDEX(Inputs!$C:$C,MATCH("*Annual Revenue Growth Rate*",Inputs!$B:$B,0)))^({col}$5-1),0))'
        )
        ws_monthly.cell(row=12, column=coluna, value=formula_revenue_l12)

        formula_expense_l16 = (
            f'=IF($B16="",0,IF(AND({col}$3>=INDEX(Inputs!$C:$C,MATCH("*Expense Start Month*",Inputs!$B:$B,0)),'
            f'{col}$3<=INDEX(Inputs!$C:$C,MATCH("End Year",Inputs!$B:$B,0))*12),'
            f'IF($B16="Property Management Fee (%EGI)",-{col}11*Inputs!$C$28,-IncomeExpenses!$C$90/12*'
            f'(1+INDEX(Inputs!$C:$C,MATCH("*Annual Expense Growth Rate*",Inputs!$B:$B,0)))^({col}$5-1)),0))'
        )
        ws_monthly.cell(row=16, column=coluna, value=formula_expense_l16)

        formula_expense_l17 = (
            f'=IF($B17="",0,IF(AND({col}$3>=INDEX(Inputs!$C:$C,MATCH("*Expense Start Month*",Inputs!$B:$B,0)),'
            f'{col}$3<=INDEX(Inputs!$C:$C,MATCH("End Year",Inputs!$B:$B,0))*12),'
            f'IF($B17="Property Management Fee (%EGI)",-{col}11*Inputs!$C$28,-IncomeExpenses!$C$90/12*'
            f'(1+INDEX(Inputs!$C:$C,MATCH("*Annual Expense Growth Rate*",Inputs!$B:$B,0)))^({col}$5-1)),0))'
        )
        ws_monthly.cell(row=17, column=coluna, value=formula_expense_l17)

        formula_total_expenses = f'=SUM({col}{row_expense_start}:{col}{row_expense_end})'
        ws_monthly.cell(row=row_total_expenses, column=coluna, value=formula_total_expenses)
        ws_monthly.cell(row=row_noi, column=coluna, value=f'={col}{row_total_revenue}+{col}{row_total_expenses}')

        # CapEx mensal usa lista de itens da Inputs (mes de execucao em coluna D).
        formula_capex = (
            f'=IF({col}$3<=INDEX(Inputs!$C:$C,MATCH("End Year",Inputs!$B:$B,0))*12,'
            f'-SUMIFS(Inputs!$C:$C,Inputs!$B:$B,"*CapEx Item*",Inputs!$D:$D,{col}$3),0)'
        )
        ws_monthly.cell(row=row_capex, column=coluna, value=formula_capex)

        ws_monthly.cell(
            row=row_purchase,
            column=coluna,
            value=(
                f'=IF({col}$3=1,'
                f'-INDEX(Inputs!$C:$C,MATCH("*Purchase Price*",Inputs!$B:$B,0)),0)'
            ),
        )
        ws_monthly.cell(
            row=row_closing,
            column=coluna,
            value=(
                f'=IF({col}$3=1,'
                f'-Inputs!C{row_input_closing_costs},0)'
            ),
        )
        ws_monthly.cell(
            row=row_immediate,
            column=coluna,
            value=(
                f'=IF({col}$3=1,'
                f'-Inputs!C{row_input_immediate_repairs},0)'
            ),
        )
        ws_monthly.cell(
            row=row_acq_fee,
            column=coluna,
            value=(
                f'=IF({col}$3=1,'
                f'-INDEX(Inputs!$C:$C,MATCH("*Purchase Price*",Inputs!$B:$B,0))*'
                f'Inputs!C{row_input_acquisition_fee_pct}'
                f',0)'
            ),
        )

        if coluna == 3:
            formula_due_diligence = (
                f'=IF({col}$3=1,-INDEX(Inputs!$D:$D,MATCH("*Due Diligence Costs*",Inputs!$B:$B,0)),0)'
            )
            formula_loan_origination = (
                f'=IF({col}$3=1,-INDEX(Inputs!$D:$D,MATCH("*Loan Origination Costs*",Inputs!$B:$B,0)),0)'
            )
        else:
            formula_due_diligence = "0"
            formula_loan_origination = "0"

        ws_monthly.cell(row=row_due_diligence, column=coluna, value=formula_due_diligence)
        ws_monthly.cell(row=row_loan_origination, column=coluna, value=formula_loan_origination)
        ws_monthly.cell(
            row=row_total_acquisition,
            column=coluna,
            value=f'=SUM({col}{row_purchase}:{col}{row_loan_origination})',
        )

        ws_monthly.cell(
            row=row_sale_proceeds,
            column=coluna,
            value=(
                f'=IF({col}$3=INDEX(Inputs!$C:$C,MATCH("End Year",Inputs!$B:$B,0))*12,'
                f'Inputs!C{row_input_selling_price},0)'
            ),
        )
        ws_monthly.cell(
            row=row_selling_costs,
            column=coluna,
            value=(
                f'=IF({col}$3=INDEX(Inputs!$C:$C,MATCH("End Year",Inputs!$B:$B,0))*12,'
                f'-Inputs!C{row_input_selling_costs},0)'
            ),
        )
        ws_monthly.cell(row=row_net_sale, column=coluna, value=f'={col}{row_sale_proceeds}+{col}{row_selling_costs}')
        ws_monthly.cell(
            row=row_cf_before_debt,
            column=coluna,
            value=f'={col}{row_noi}+{col}{row_capex}+{col}{row_total_acquisition}+{col}{row_net_sale}',
        )

        # Bloco Seller Finance.
        sf_end_year = 'INDEX(Inputs!$C:$C,MATCH("End Year",Inputs!$B:$B,0))'
        sf_amount = 'INDEX(Inputs!$C:$C,MATCH("*Seller Finance Amount*",Inputs!$B:$B,0))'
        sf_rate = 'INDEX(Inputs!$C:$C,MATCH("*Seller Finance Interest Rate*",Inputs!$B:$B,0))'
        sf_amort_years = 'INDEX(Inputs!$C:$C,MATCH("*Seller Finance Amortization*",Inputs!$B:$B,0))'
        sf_balloon_eoy = 'INDEX(Inputs!$C:$C,MATCH("*Seller Finance Balloon*",Inputs!$B:$B,0))'
        sf_io_period = 'INDEX(Inputs!$C:$C,MATCH("*Seller Finance IO Period*",Inputs!$B:$B,0))'
        sf_orig_month = 'INDEX(Inputs!$C:$C,MATCH("*Seller Finance Origination Month*",Inputs!$B:$B,0))'
        sf_pmt_start = 'INDEX(Inputs!$C:$C,MATCH("*Seller Finance PMT Start Month*",Inputs!$B:$B,0))'

        if coluna == 3:
            formula_sf_beginning_balance = "0"
        else:
            formula_sf_beginning_balance = (
                f'=IF({col}$3<={sf_end_year}*12,IF({col}$3<={sf_orig_month},0,{col_anterior}{row_sf_end}),0)'
            )

        formula_sf_loan_funding = f'=IF({col}$3<={sf_end_year}*12,IF({col}$3={sf_orig_month},{sf_amount},0),0)'
        formula_sf_interest_payment = (
            f'=IF({col}$3<={sf_end_year}*12,IF(OR({col}{row_sf_begin}=0,{col}$3<{sf_pmt_start}),0,-{col}{row_sf_begin}*{sf_rate}/12),0)'
        )
        formula_sf_principal_payment = (
            f'=IF({col}$3<={sf_end_year}*12,IF(OR({col}{row_sf_begin}=0,{col}$3<{sf_pmt_start}),0,'
            f'IF({col}$3-{sf_pmt_start}+1<={sf_io_period},0,MIN(PMT({sf_rate}/12,{sf_amort_years}*12,-{sf_amount})-{col}{row_sf_begin}*{sf_rate}/12,{col}{row_sf_begin}))),0)'
        )
        formula_sf_loan_payoff = (
            f'=IF({col}$3<={sf_end_year}*12,IF(OR({col}$3={sf_balloon_eoy}*12,{col}$3={sf_end_year}*12),-MAX({col}{row_sf_begin}+{col}{row_sf_funding}-{col}{row_sf_principal},0),0),0)'
        )
        formula_sf_ending_balance = (
            f'=IF({col}$3<={sf_end_year}*12,MAX({col}{row_sf_begin}+{col}{row_sf_funding}-{col}{row_sf_principal}+{col}{row_sf_payoff},0),0)'
        )
        formula_sf_total_debt_service = (
            f'=IF({col}$3<={sf_end_year}*12,{col}{row_sf_interest}-{col}{row_sf_principal}+{col}{row_sf_payoff},0)'
        )

        ws_monthly.cell(row=row_sf_begin, column=coluna, value=formula_sf_beginning_balance)
        ws_monthly.cell(row=row_sf_funding, column=coluna, value=formula_sf_loan_funding)
        ws_monthly.cell(row=row_sf_interest, column=coluna, value=formula_sf_interest_payment)
        ws_monthly.cell(row=row_sf_principal, column=coluna, value=formula_sf_principal_payment)
        ws_monthly.cell(row=row_sf_payoff, column=coluna, value=formula_sf_loan_payoff)
        ws_monthly.cell(row=row_sf_end, column=coluna, value=formula_sf_ending_balance)
        ws_monthly.cell(row=row_sf_total, column=coluna, value=formula_sf_total_debt_service)

        # Bloco Bank Loan.
        bl_end_year = 'INDEX(Inputs!$C:$C,MATCH("End Year",Inputs!$B:$B,0))'
        bl_amount = 'INDEX(Inputs!$C:$C,MATCH("*Bank Loan Amount*",Inputs!$B:$B,0))'
        bl_rate = 'INDEX(Inputs!$C:$C,MATCH("*Bank Loan Interest Rate*",Inputs!$B:$B,0))'
        bl_amort_years = 'INDEX(Inputs!$C:$C,MATCH("*Bank Loan Amortization*",Inputs!$B:$B,0))'
        bl_balloon_eoy = 'INDEX(Inputs!$C:$C,MATCH("*Bank Loan Balloon*",Inputs!$B:$B,0))'
        bl_io_period = 'INDEX(Inputs!$C:$C,MATCH("*Bank Loan IO Period*",Inputs!$B:$B,0))'
        bl_orig_month = 'INDEX(Inputs!$C:$C,MATCH("*Bank Loan Origination Month*",Inputs!$B:$B,0))'
        bl_pmt_start = 'INDEX(Inputs!$C:$C,MATCH("*Bank Loan PMT Start Month*",Inputs!$B:$B,0))'

        if coluna == 3:
            formula_bl_beginning_balance = "0"
        else:
            formula_bl_beginning_balance = (
                f'=IF({col}$3<={bl_end_year}*12,IF({col}$3<={bl_orig_month},0,{col_anterior}{row_bl_end}),0)'
            )

        formula_bl_loan_funding = f'=IF({col}$3<={bl_end_year}*12,IF({col}$3={bl_orig_month},{bl_amount},0),0)'
        formula_bl_interest_payment = (
            f'=IF({col}$3<={bl_end_year}*12,IF(OR({col}{row_bl_begin}=0,{col}$3<{bl_pmt_start}),0,-{col}{row_bl_begin}*{bl_rate}/12),0)'
        )
        formula_bl_principal_payment = (
            f'=IF({col}$3<={bl_end_year}*12,IF(OR({col}{row_bl_begin}=0,{col}$3<{bl_pmt_start}),0,'
            f'IF({col}$3-{bl_pmt_start}+1<={bl_io_period},0,MIN(PMT({bl_rate}/12,{bl_amort_years}*12,-{bl_amount})-{col}{row_bl_begin}*{bl_rate}/12,{col}{row_bl_begin}))),0)'
        )
        formula_bl_loan_payoff = (
            f'=IF({col}$3<={bl_end_year}*12,IF(OR({col}$3={bl_balloon_eoy}*12,{col}$3={bl_end_year}*12),-MAX({col}{row_bl_begin}+{col}{row_bl_funding}-{col}{row_bl_principal},0),0),0)'
        )
        formula_bl_ending_balance = (
            f'=IF({col}$3<={bl_end_year}*12,MAX({col}{row_bl_begin}+{col}{row_bl_funding}-{col}{row_bl_principal}+{col}{row_bl_payoff},0),0)'
        )
        formula_bl_total_debt_service = (
            f'=IF({col}$3<={bl_end_year}*12,{col}{row_bl_interest}-{col}{row_bl_principal}+{col}{row_bl_payoff},0)'
        )

        ws_monthly.cell(row=row_bl_begin, column=coluna, value=formula_bl_beginning_balance)
        ws_monthly.cell(row=row_bl_funding, column=coluna, value=formula_bl_loan_funding)
        ws_monthly.cell(row=row_bl_interest, column=coluna, value=formula_bl_interest_payment)
        ws_monthly.cell(row=row_bl_principal, column=coluna, value=formula_bl_principal_payment)
        ws_monthly.cell(row=row_bl_payoff, column=coluna, value=formula_bl_loan_payoff)
        ws_monthly.cell(row=row_bl_end, column=coluna, value=formula_bl_ending_balance)
        ws_monthly.cell(row=row_bl_total, column=coluna, value=formula_bl_total_debt_service)

        # Bloco Refinance Loan.
        rf_end_year = 'INDEX(Inputs!$C:$C,MATCH("End Year",Inputs!$B:$B,0))'
        rf_year = 'INDEX(Inputs!$C:$C,MATCH("Refi Year",Inputs!$B:$B,0))'
        rf_amount = 'INDEX(Inputs!$C:$C,MATCH("*Refi Loan Amount*",Inputs!$B:$B,0))'
        rf_rate = 'INDEX(Inputs!$C:$C,MATCH("*Refi Interest Rate*",Inputs!$B:$B,0))'
        rf_amort_years = 'INDEX(Inputs!$C:$C,MATCH("*Refi Amortization*",Inputs!$B:$B,0))'
        rf_term_years = 'INDEX(Inputs!$C:$C,MATCH("*Refi Term*",Inputs!$B:$B,0))'
        rf_io_period = 'INDEX(Inputs!$C:$C,MATCH("*Refi IO Period*",Inputs!$B:$B,0))'

        if coluna == 3:
            formula_rf_beginning_balance = "0"
        else:
            formula_rf_beginning_balance = (
                f'=IF({col}$3<={rf_end_year}*12,IF({col}$3<={rf_year}*12,0,{col_anterior}{row_rf_end}),0)'
            )

        formula_rf_loan_funding = f'=IF({col}$3<={rf_end_year}*12,IF({col}$3={rf_year}*12+1,{rf_amount},0),0)'
        formula_rf_interest_payment = f'=IF({col}$3<={rf_end_year}*12,IF({col}{row_rf_begin}=0,0,-{col}{row_rf_begin}*{rf_rate}/12),0)'
        formula_rf_principal_payment = (
            f'=IF({col}$3<={rf_end_year}*12,IF({col}{row_rf_begin}=0,0,'
            f'IF({col}$3-{rf_year}*12<={rf_io_period},0,MIN(IFERROR(PMT({rf_rate}/12,{rf_amort_years}*12,-{rf_amount}),0)-{col}{row_rf_begin}*{rf_rate}/12,{col}{row_rf_begin}))),0)'
        )
        formula_rf_loan_payoff = (
            f'=IF({col}$3<={rf_end_year}*12,IF(OR({col}$3=({rf_year}+{rf_term_years})*12,{col}$3={rf_end_year}*12),-MAX({col}{row_rf_begin}+{col}{row_rf_funding}-{col}{row_rf_principal},0),0),0)'
        )
        formula_rf_ending_balance = (
            f'=IF({col}$3<={rf_end_year}*12,MAX({col}{row_rf_begin}+{col}{row_rf_funding}-{col}{row_rf_principal}+{col}{row_rf_payoff},0),0)'
        )
        formula_rf_total_debt_service = (
            f'=IF({col}$3<={rf_end_year}*12,{col}{row_rf_interest}-{col}{row_rf_principal}+{col}{row_rf_payoff},0)'
        )

        ws_monthly.cell(row=row_rf_begin, column=coluna, value=formula_rf_beginning_balance)
        ws_monthly.cell(row=row_rf_funding, column=coluna, value=formula_rf_loan_funding)
        ws_monthly.cell(row=row_rf_interest, column=coluna, value=formula_rf_interest_payment)
        ws_monthly.cell(row=row_rf_principal, column=coluna, value=formula_rf_principal_payment)
        ws_monthly.cell(row=row_rf_payoff, column=coluna, value=formula_rf_loan_payoff)
        ws_monthly.cell(row=row_rf_end, column=coluna, value=formula_rf_ending_balance)
        ws_monthly.cell(row=row_rf_total, column=coluna, value=formula_rf_total_debt_service)

        # Equity, net refi, loan proceeds e cashflow apos divida.
        ws_monthly.cell(
            row=row_equity,
            column=coluna,
            value=(
                f'=IF({col}$3=1,'
                f'Inputs!C{row_input_equity_contribution},0)'
            ),
        )
        ws_monthly.cell(
            row=row_refi_net,
            column=coluna,
            value=(
                f'=IF({col}$3<=INDEX(Inputs!$C:$C,MATCH("End Year",Inputs!$B:$B,0))*12,'
                f'IF({col}$3=INDEX(Inputs!$C:$C,MATCH("Refi Year",Inputs!$B:$B,0))*12+1,'
                f'{col}{row_rf_funding}+{col}{row_bl_payoff}-INDEX(Inputs!$C:$C,MATCH("*Refi Closing Costs ($)*",Inputs!$B:$B,0)),0),0)'
            ),
        )
        ws_monthly.cell(row=row_loan_proceeds, column=coluna, value=f'=IF({col}$3<=INDEX(Inputs!$C:$C,MATCH("End Year",Inputs!$B:$B,0))*12,{col}{row_sf_funding}+{col}{row_bl_funding},0)')
        ws_monthly.cell(row=row_total_debt, column=coluna, value=f'={col}{row_sf_total}+{col}{row_bl_total}+{col}{row_rf_total}')
        ws_monthly.cell(
            row=row_cf_after_debt,
            column=coluna,
            value=f'={col}{row_cf_before_debt}+{col}{row_total_debt}+{col}{row_equity}+{col}{row_refi_net}+{col}{row_loan_proceeds}',
        )

        if coluna == 3:
            ws_monthly.cell(row=row_cum_cf, column=coluna, value=f'={col}{row_cf_after_debt}')
        else:
            ws_monthly.cell(
                row=row_cum_cf,
                column=coluna,
                value=f'={col_anterior}{row_cum_cf}+{col}{row_cf_after_debt}',
            )

        # Distribuicoes para membros (LP/GP) e metricas acumuladas.
        ws_monthly.cell(
            row=row_asset_fee,
            column=coluna,
            value=(
                f'=IF({col}$3<=INDEX(Inputs!$C:$C,MATCH("End Year",Inputs!$B:$B,0))*12,'
                f'IF({col}{row_cf_before_debt}+{col}{row_total_debt}<=0,0,-{col}{row_egi}*'
                f'Inputs!C{row_input_asset_mgmt_fee}),0)'
            ),
        )
        ws_monthly.cell(
            row=row_cash_members,
            column=coluna,
            value=f'=MAX({col}{row_cf_before_debt}+{col}{row_total_debt}+{col}{row_asset_fee}-{col}{row_return_capital},0)',
        )
        ws_monthly.cell(
            row=row_cap_partner_dist,
            column=coluna,
            value=f'={col}{row_cash_members}*Inputs!C{row_input_cap_partner_share}',
        )
        ws_monthly.cell(
            row=row_return_capital,
            column=coluna,
            value=(
                f'=IF({col}$3=INDEX(Inputs!$C:$C,MATCH("End Year",Inputs!$B:$B,0))*12,'
                f'Inputs!C{row_input_cap_partner_contribution},0)'
            ),
        )
        ws_monthly.cell(
            row=row_manager_dist,
            column=coluna,
            value=f'={col}{row_cash_members}*Inputs!C{row_input_manager_share}',
        )
        ws_monthly.cell(row=row_total_dist, column=coluna, value=f'={col}{row_cap_partner_dist}+{col}{row_return_capital}+{col}{row_manager_dist}')

        if coluna == 3:
            ws_monthly.cell(row=row_cum_noi, column=coluna, value=f'={col}{row_noi}')
            ws_monthly.cell(row=row_cum_cf_after_debt, column=coluna, value=f'={col}{row_cf_after_debt}')
        else:
            ws_monthly.cell(
                row=row_cum_noi,
                column=coluna,
                value=f'={col_anterior}{row_cum_noi}+{col}{row_noi}',
            )
            ws_monthly.cell(
                row=row_cum_cf_after_debt,
                column=coluna,
                value=f'={col_anterior}{row_cum_cf_after_debt}+{col}{row_cf_after_debt}',
            )

        ws_monthly.cell(
            row=row_dscr,
            column=coluna,
            value=f'=IF({col}{row_total_debt}=0,"-",{col}{row_noi}/ABS({col}{row_total_debt}))',
        )


def _aplicar_formulas_resumo_cf(ws_resumo, ws_monthly, periodicidade: str) -> None:
    """Aplica formulas de agregacao no Quarterly CF ou Annual CF.

Onde atua:
- Destino: aba Quarterly CF ou Annual CF.
- Origem: aba Monthly CF.

O que faz:
- Soma linhas mensais por trimestre/ano via SUMIF.
- Traz saldos inicial/final de divida por INDEX+MATCH no mes correto.
- Replica distribuicoes e metricas no nivel agregado.

Por que faz:
- Garante consistencia entre visao mensal e visoes consolidadas.
    """
    layout_r = _obter_layout_resumo_cf(ws_resumo)
    layout_m = _obter_layout_monthly_cf(ws_monthly)

    crit_row = 4 if periodicidade == "quarterly" else 5
    fator_periodo = 3 if periodicidade == "quarterly" else 12

    income_rows_monthly = list(range(layout_m["income_start"], layout_m["income_end"] + 1))
    expense_rows_monthly = list(range(layout_m["expense_start"], layout_m["expense_end"] + 1))

    for idx, linha_monthly in enumerate(income_rows_monthly):
        linha_resumo = layout_r["income_start"] + idx
        ws_resumo.cell(row=linha_resumo, column=2, value=ws_monthly.cell(row=linha_monthly, column=2).value)

    for idx, linha_monthly in enumerate(expense_rows_monthly):
        linha_resumo = layout_r["expense_start"] + idx
        ws_resumo.cell(row=linha_resumo, column=2, value=ws_monthly.cell(row=linha_monthly, column=2).value)

    for coluna in range(3, ws_resumo.max_column + 1):
        col = get_column_letter(coluna)
        col_anterior = get_column_letter(coluna - 1) if coluna > 3 else None

        # Helper local: agrega uma linha do Monthly CF para o periodo do resumo.
        def _sumif_row(linha_monthly: int) -> str:
            return (
                f"=SUMIF('Monthly CF'!$C${crit_row}:$DR${crit_row},{col}$3,'Monthly CF'!$C${linha_monthly}:$DR${linha_monthly})"
            )

        ws_resumo.cell(row=layout_r["row_gross"], column=coluna, value=_sumif_row(8))
        ws_resumo.cell(row=layout_r["row_vacancy"], column=coluna, value=_sumif_row(9))
        ws_resumo.cell(row=layout_r["row_credit_loss"], column=coluna, value=_sumif_row(10))
        ws_resumo.cell(row=layout_r["row_egi"], column=coluna, value=_sumif_row(11))

        for idx, linha_monthly in enumerate(income_rows_monthly):
            linha_resumo = layout_r["income_start"] + idx
            ws_resumo.cell(row=linha_resumo, column=coluna, value=_sumif_row(linha_monthly))

        ws_resumo.cell(row=layout_r["row_total_revenue"], column=coluna, value=_sumif_row(layout_m["row_total_revenue"]))

        for idx, linha_monthly in enumerate(expense_rows_monthly):
            linha_resumo = layout_r["expense_start"] + idx
            ws_resumo.cell(row=linha_resumo, column=coluna, value=_sumif_row(linha_monthly))

        ws_resumo.cell(row=layout_r["row_total_expenses"], column=coluna, value=_sumif_row(layout_m["row_total_expenses"]))
        ws_resumo.cell(row=layout_r["row_noi"], column=coluna, value=_sumif_row(layout_m["row_noi"]))
        ws_resumo.cell(row=layout_r["row_capex"], column=coluna, value=_sumif_row(layout_m["row_capex"]))
        ws_resumo.cell(row=layout_r["row_acquisition"], column=coluna, value=_sumif_row(layout_m["row_total_acquisition"]))
        ws_resumo.cell(row=layout_r["row_sale"], column=coluna, value=_sumif_row(layout_m["row_sale_proceeds"]))
        ws_resumo.cell(row=layout_r["row_selling"], column=coluna, value=_sumif_row(layout_m["row_selling_costs"]))
        ws_resumo.cell(row=layout_r["row_net_sale"], column=coluna, value=_sumif_row(layout_m["row_net_sale"]))
        ws_resumo.cell(row=layout_r["row_cf_before_debt"], column=coluna, value=_sumif_row(layout_m["row_cf_before_debt"]))

        ws_resumo.cell(
            row=layout_r["row_sf_begin"],
            column=coluna,
            value=(
                f"=INDEX('Monthly CF'!$C${layout_m['row_sf_begin']}:$DR${layout_m['row_sf_begin']},"
                f"MATCH(({col}$3-1)*{fator_periodo}+1,'Monthly CF'!$C$3:$DR$3,0))"
            ),
        )
        ws_resumo.cell(row=layout_r["row_sf_funding"], column=coluna, value=_sumif_row(layout_m["row_sf_funding"]))
        ws_resumo.cell(row=layout_r["row_sf_interest"], column=coluna, value=_sumif_row(layout_m["row_sf_interest"]))
        ws_resumo.cell(row=layout_r["row_sf_principal"], column=coluna, value=_sumif_row(layout_m["row_sf_principal"]))
        ws_resumo.cell(row=layout_r["row_sf_payoff"], column=coluna, value=_sumif_row(layout_m["row_sf_payoff"]))
        ws_resumo.cell(
            row=layout_r["row_sf_end"],
            column=coluna,
            value=(
                f"=INDEX('Monthly CF'!$C${layout_m['row_sf_end']}:$DR${layout_m['row_sf_end']},"
                f"MATCH({col}$3*{fator_periodo},'Monthly CF'!$C$3:$DR$3,0))"
            ),
        )
        ws_resumo.cell(row=layout_r["row_sf_total"], column=coluna, value=_sumif_row(layout_m["row_sf_total"]))

        ws_resumo.cell(
            row=layout_r["row_bl_begin"],
            column=coluna,
            value=(
                f"=INDEX('Monthly CF'!$C${layout_m['row_bl_begin']}:$DR${layout_m['row_bl_begin']},"
                f"MATCH(({col}$3-1)*{fator_periodo}+1,'Monthly CF'!$C$3:$DR$3,0))"
            ),
        )
        ws_resumo.cell(row=layout_r["row_bl_funding"], column=coluna, value=_sumif_row(layout_m["row_bl_funding"]))
        ws_resumo.cell(row=layout_r["row_bl_interest"], column=coluna, value=_sumif_row(layout_m["row_bl_interest"]))
        ws_resumo.cell(row=layout_r["row_bl_principal"], column=coluna, value=_sumif_row(layout_m["row_bl_principal"]))
        ws_resumo.cell(row=layout_r["row_bl_payoff"], column=coluna, value=_sumif_row(layout_m["row_bl_payoff"]))
        ws_resumo.cell(
            row=layout_r["row_bl_end"],
            column=coluna,
            value=(
                f"=INDEX('Monthly CF'!$C${layout_m['row_bl_end']}:$DR${layout_m['row_bl_end']},"
                f"MATCH({col}$3*{fator_periodo},'Monthly CF'!$C$3:$DR$3,0))"
            ),
        )
        ws_resumo.cell(row=layout_r["row_bl_total"], column=coluna, value=_sumif_row(layout_m["row_bl_total"]))

        ws_resumo.cell(
            row=layout_r["row_rf_begin"],
            column=coluna,
            value=(
                f"=INDEX('Monthly CF'!$C${layout_m['row_rf_begin']}:$DR${layout_m['row_rf_begin']},"
                f"MATCH(({col}$3-1)*{fator_periodo}+1,'Monthly CF'!$C$3:$DR$3,0))"
            ),
        )
        ws_resumo.cell(row=layout_r["row_rf_funding"], column=coluna, value=_sumif_row(layout_m["row_rf_funding"]))
        ws_resumo.cell(row=layout_r["row_rf_interest"], column=coluna, value=_sumif_row(layout_m["row_rf_interest"]))
        ws_resumo.cell(row=layout_r["row_rf_principal"], column=coluna, value=_sumif_row(layout_m["row_rf_principal"]))
        ws_resumo.cell(row=layout_r["row_rf_payoff"], column=coluna, value=_sumif_row(layout_m["row_rf_payoff"]))
        ws_resumo.cell(
            row=layout_r["row_rf_end"],
            column=coluna,
            value=(
                f"=INDEX('Monthly CF'!$C${layout_m['row_rf_end']}:$DR${layout_m['row_rf_end']},"
                f"MATCH({col}$3*{fator_periodo},'Monthly CF'!$C$3:$DR$3,0))"
            ),
        )
        ws_resumo.cell(row=layout_r["row_rf_total"], column=coluna, value=_sumif_row(layout_m["row_rf_total"]))

        ws_resumo.cell(row=layout_r["row_equity"], column=coluna, value=_sumif_row(layout_m["row_equity"]))
        ws_resumo.cell(row=layout_r["row_refi_net"], column=coluna, value=_sumif_row(layout_m["row_refi_net"]))
        ws_resumo.cell(row=layout_r["row_loan_proceeds"], column=coluna, value=_sumif_row(layout_m["row_loan_proceeds"]))
        ws_resumo.cell(row=layout_r["row_total_debt"], column=coluna, value=_sumif_row(layout_m["row_total_debt"]))
        ws_resumo.cell(row=layout_r["row_cf_after_debt"], column=coluna, value=_sumif_row(layout_m["row_cf_after_debt"]))

        if coluna == 3:
            ws_resumo.cell(row=layout_r["row_cum_cf"], column=coluna, value=f'={col}{layout_r["row_cf_after_debt"]}')
        else:
            ws_resumo.cell(
                row=layout_r["row_cum_cf"],
                column=coluna,
                value=f'={col_anterior}{layout_r["row_cum_cf"]}+{col}{layout_r["row_cf_after_debt"]}',
            )

        ws_resumo.cell(row=layout_r["row_asset_fee"], column=coluna, value=_sumif_row(layout_m["row_asset_fee"]))
        ws_resumo.cell(row=layout_r["row_cash_members"], column=coluna, value=_sumif_row(layout_m["row_cash_members"]))
        ws_resumo.cell(row=layout_r["row_cap_partner_dist"], column=coluna, value=_sumif_row(layout_m["row_cap_partner_dist"]))
        ws_resumo.cell(row=layout_r["row_return_capital"], column=coluna, value=_sumif_row(layout_m["row_return_capital"]))
        ws_resumo.cell(row=layout_r["row_manager_dist"], column=coluna, value=_sumif_row(layout_m["row_manager_dist"]))
        ws_resumo.cell(row=layout_r["row_total_dist"], column=coluna, value=_sumif_row(layout_m["row_total_dist"]))
        ws_resumo.cell(
            row=layout_r["row_dscr"],
            column=coluna,
            value=f'=IF(ABS({col}{layout_r["row_total_debt"]})=0,"-",{col}{layout_r["row_noi"]}/ABS({col}{layout_r["row_total_debt"]}))',
        )


def _atualizar_labels_resumo_cf(wb) -> None:
    """Sincroniza labels e formulas das abas de resumo com a Monthly CF.

Onde atua:
- Origem: Monthly CF.
- Destino: Quarterly CF e Annual CF.

Por que faz:
- Quando o bloco dinamico muda na Monthly, os resumos precisam espelhar a mesma estrutura.
    """
    if "Monthly CF" not in wb.sheetnames:
        return

    ws_monthly = wb["Monthly CF"]

    if "Quarterly CF" in wb.sheetnames:
        _aplicar_formulas_resumo_cf(wb["Quarterly CF"], ws_monthly, periodicidade="quarterly")

    if "Annual CF" in wb.sheetnames:
        _aplicar_formulas_resumo_cf(wb["Annual CF"], ws_monthly, periodicidade="annual")


def _atualizar_formula_summary_exit_cap_rate(wb, ws_inputs) -> None:
    """Atualiza no Summary a formula de Exit Cap Rate apontando para Inputs.

Onde atua:
- Aba Summary, celula C14.
- Aba Inputs para descobrir a linha correta do label "Exit Cap Rate (%)".

Por que faz:
- Evita que C14 fique preso a uma linha antiga quando a Inputs muda.
    """
    if "Summary" not in wb.sheetnames:
        return

    row_exit_cap_rate = _obter_linha_input_obrigatoria(
        ws_inputs,
        "Exit Cap Rate (%)",
        correspondencia_exata=True,
        ultima_ocorrencia=True,
    )
    _set_cell_value_respeitando_merge(wb["Summary"], "C14", f"=Inputs!C{row_exit_cap_rate}")


def _atualizar_formulas_equity_waterfall(wb) -> None:
    """Reconstrui formulas da aba Equity Waterfall com referencias dinamicas.

Onde atua:
- Destino: aba Equity Waterfall.
- Origem: Inputs (parametros de contribuicao/split/hurdles) e Annual CF (cashflow anual).

Por que faz:
- O Waterfall depende de varios labels sensiveis; sem busca dinamica por label,
  qualquer deslocamento na Inputs quebra os calculos de tiers e IRR.
    """
    if "Equity Waterfall" not in wb.sheetnames or "Inputs" not in wb.sheetnames:
        return

    ws_eq = wb["Equity Waterfall"]
    ws_inputs = wb["Inputs"]

    row_equity_contribution = _obter_linha_input_obrigatoria(ws_inputs, "Equity Contribution", correspondencia_exata=True, ultima_ocorrencia=True)
    row_cap_partner_contribution = _obter_linha_input_obrigatoria(ws_inputs, "Capital Partner Contribution ($)", correspondencia_exata=True, ultima_ocorrencia=True)
    row_manager_contribution = _obter_linha_input_obrigatoria(ws_inputs, "Manager Contribution ($)", correspondencia_exata=True, ultima_ocorrencia=True)
    row_cap_partner_share = _obter_linha_input_obrigatoria(ws_inputs, "Capital Partner Share (%)", correspondencia_exata=True, ultima_ocorrencia=True)
    row_manager_share = _obter_linha_input_obrigatoria(ws_inputs, "Manager Share (%)", correspondencia_exata=True, ultima_ocorrencia=True)
    row_pref_return_annual = _obter_linha_input_obrigatoria(ws_inputs, "Preferred Return (Annual)", correspondencia_exata=True, ultima_ocorrencia=True)
    row_gp_catchup_target = _obter_linha_input_obrigatoria(ws_inputs, "GP Catch-Up Target (%)", correspondencia_exata=True, ultima_ocorrencia=True)
    row_gp_catchup_share = _obter_linha_input_obrigatoria(ws_inputs, "GP Catch-Up Share (%)", correspondencia_exata=True, ultima_ocorrencia=True)
    row_irr_hurdle = _obter_linha_input_obrigatoria(ws_inputs, "IRR Hurdle", correspondencia_exata=True, ultima_ocorrencia=True)
    row_tier_1 = _obter_linha_input_obrigatoria(ws_inputs, "1. Return of Capital", correspondencia_exata=True, ultima_ocorrencia=True)
    row_tier_2 = _obter_linha_input_obrigatoria(ws_inputs, "2. Preferred Return", correspondencia_exata=True, ultima_ocorrencia=True)
    row_tier_3 = _obter_linha_input_obrigatoria(ws_inputs, "3. GP Catch-Up", correspondencia_exata=True, ultima_ocorrencia=True)
    row_tier_4 = _obter_linha_input_obrigatoria(ws_inputs, "4. Residual (below IRR hurdle)", correspondencia_exata=True, ultima_ocorrencia=True)
    row_tier_5 = _obter_linha_input_obrigatoria(ws_inputs, "5. Residual (above IRR hurdle)", correspondencia_exata=True, ultima_ocorrencia=True)

    ws_eq["C6"] = f"=-Inputs!C{row_equity_contribution}"
    ws_eq["C12"] = f"=-Inputs!C{row_cap_partner_contribution}"
    ws_eq["C13"] = f"=-Inputs!C{row_manager_contribution}"
    ws_eq["C15"] = f"=Inputs!C{row_cap_partner_share}"
    ws_eq["C16"] = f"=Inputs!C{row_manager_share}"

    ws_eq["B21"] = f'="Tier 1: Return of Capital ("&TEXT(Inputs!$C${row_tier_1},"0%")&" to LP)"'
    ws_eq["B27"] = f'="Tier 2: Preferred Return ("&TEXT(Inputs!$C${row_tier_2},"0%")&" to LP)"'
    ws_eq["B34"] = f'="Tier 3: GP Catch-Up ("&TEXT(Inputs!$C${row_tier_3},"0%")&"/"&TEXT(Inputs!$D${row_tier_3},"0%")&")"'
    ws_eq["B40"] = f'="Tier 4: Residual Split (below "&TEXT(Inputs!$C${row_irr_hurdle},"0%")&" IRR)"'
    ws_eq["B41"] = f'="  Tier 4 to LP ("&TEXT(Inputs!$C${row_tier_4},"0%")&")"'
    ws_eq["B42"] = f'="  Tier 4 to GP ("&TEXT(Inputs!$D${row_tier_4},"0%")&")"'
    ws_eq["B45"] = f'="Tier 5: Residual Split (above "&TEXT(Inputs!$C${row_irr_hurdle},"0%")&" IRR)"'
    ws_eq["B46"] = f'="  Tier 5 to LP ("&TEXT(Inputs!$C${row_tier_5},"0%")&")"'
    ws_eq["B47"] = f'="  Tier 5 to GP ("&TEXT(Inputs!$D${row_tier_5},"0%")&")"'

    for coluna in range(3, 14):
        col = get_column_letter(coluna)
        ws_eq.cell(row=29, column=coluna, value=f"={col}22*Inputs!$C${row_pref_return_annual}")

    for linha_zero in (35, 36, 37, 41, 42, 46, 47):
        ws_eq.cell(row=linha_zero, column=3, value=0)

    for coluna in range(4, 14):
        col = get_column_letter(coluna)
        col_prev = get_column_letter(coluna - 1)
        irr_range = f"C$57" if coluna == 4 else f"C$57:{col_prev}$57"

        ws_eq.cell(
            row=35,
            column=coluna,
            value=f"=MIN({col}32,{col}30*Inputs!$C${row_gp_catchup_target}/(1-Inputs!$C${row_gp_catchup_target}))",
        )
        ws_eq.cell(row=36, column=coluna, value=f"={col}35*(1-Inputs!$C${row_gp_catchup_share})")
        ws_eq.cell(row=37, column=coluna, value=f"={col}35*Inputs!$C${row_gp_catchup_share}")

        ws_eq.cell(
            row=41,
            column=coluna,
            value=f"=IF(IFERROR(IRR({irr_range}),0)<Inputs!$C${row_irr_hurdle},{col}38*Inputs!$C${row_tier_4},0)",
        )
        ws_eq.cell(
            row=42,
            column=coluna,
            value=f"=IF(IFERROR(IRR({irr_range}),0)<Inputs!$C${row_irr_hurdle},{col}38*Inputs!$D${row_tier_4},0)",
        )
        ws_eq.cell(
            row=46,
            column=coluna,
            value=f"=IF(IFERROR(IRR({irr_range}),0)>=Inputs!$C${row_irr_hurdle},{col}38*Inputs!$C${row_tier_5},{col}43*Inputs!$C${row_tier_5})",
        )
        ws_eq.cell(
            row=47,
            column=coluna,
            value=f"=IF(IFERROR(IRR({irr_range}),0)>=Inputs!$C${row_irr_hurdle},{col}38*Inputs!$D${row_tier_5},{col}43*Inputs!$D${row_tier_5})",
        )

    if "Annual CF" not in wb.sheetnames:
        return

    ws_annual = wb["Annual CF"]
    try:
        row_cf_after_debt = _obter_layout_resumo_cf(ws_annual)["row_cf_after_debt"]
    except Exception:
        row_cf_after_debt = _encontrar_linha_por_texto(ws_annual, "cashflow after debt service", 1)

    if not row_cf_after_debt:
        return

    max_col = ws_eq.max_column
    if max_col < 4:
        return

    ws_eq.cell(row=7, column=3, value=0)
    for coluna in range(4, max_col):
        col_annual = get_column_letter(coluna - 1)
        ws_eq.cell(row=7, column=coluna, value=f"='Annual CF'!{col_annual}{row_cf_after_debt}")

    col_penultima = get_column_letter(max_col - 1)
    ws_eq.cell(row=7, column=max_col, value=f"=SUM(C7:{col_penultima}7)")


def recalcular_formulas_proforma_inputs(ws_inputs, ws_ie) -> None:
    """Recalcula formulas de Proforma nas abas Inputs e IncomeExpenses.

Onde atua:
- Inputs: coluna E para campos com crescimento anual ou percentuais diretos.
- IncomeExpenses: coluna E para projeções anuais de renda/despesa.

Por que faz:
- Mantem a base de crescimento usada depois pela Monthly CF sempre coerente
    com os labels realmente preenchidos pelo usuario.
        """
    for linha in range(1, ws_inputs.max_row + 1):
        valor_b = ws_inputs.cell(row=linha, column=2).value
        if not isinstance(valor_b, str):
            continue

        nome = valor_b.strip().lower()

        if "annual revenue growth rate" in nome or "revenue start month" in nome:
            ws_inputs.cell(row=linha, column=5, value=None)
            continue

        if "annual expense growth rate" in nome or "expense start month" in nome:
            ws_inputs.cell(row=linha, column=5, value=None)
            continue

        if "vacancy" in nome or "credit loss" in nome or "property management fee" in nome:
            ws_inputs.cell(row=linha, column=5, value=f"=C{linha}")
            continue

        if "gross potential rent" in nome:
            ws_inputs.cell(row=linha, column=5, value=f"=C{linha}*(1+D{linha})")

    for linha in range(IE_INCOME_START, IE_INCOME_END + 1):
        nome_income = ws_ie.cell(row=linha, column=2).value
        if isinstance(nome_income, str) and nome_income.strip():
            ws_ie.cell(row=linha, column=5, value=f"=C{linha}*(1+D{linha})")
        else:
            ws_ie.cell(row=linha, column=5, value=None)

    for linha in range(IE_EXPENSE_START, IE_EXPENSE_END + 1):
        nome_expense = ws_ie.cell(row=linha, column=2).value
        if not isinstance(nome_expense, str) or not nome_expense.strip():
            ws_ie.cell(row=linha, column=5, value=None)
            continue

        if "management fee" in nome_expense.strip().lower():
            ws_ie.cell(row=linha, column=5, value=f"=C{linha}")
        else:
            ws_ie.cell(row=linha, column=5, value=f"=C{linha}*(1+D{linha})")


def _recalcular_formulas_estrutura_inputs(ws_inputs, ws_monthly=None) -> None:
    """Reconstrui formulas estruturais da Inputs por secao e label.

Onde atua:
- Aba Inputs (Sources & Uses, Refi assumptions, Exit values, Waterfall setup).
- Opcionalmente usa Monthly CF para calcular NOI por ano sem linha fixa.

O que faz:
- Identifica limites de secoes (topo, uses, parametros, exit values, waterfall).
- Reaplica formulas de compra, custos, contribuicoes, refi e venda.
- Reaplica splits de tiers no bloco de waterfall.

Por que faz:
- Inputs costuma ter labels repetidos; ancorar por secao evita colisoes de match.
    """

    max_row = ws_inputs.max_row

    def _label(linha: int) -> str:
        return _texto_normalizado(ws_inputs.cell(row=linha, column=2).value)

    def _limite_final(end: int | None) -> int:
        if end is None:
            return max_row
        return min(max_row, max(1, end))

    def _first_contains(texto: str, start: int = 1, end: int | None = None) -> int | None:
        alvo = _texto_normalizado(texto)
        fim = _limite_final(end)
        for linha in range(max(1, start), fim + 1):
            if alvo in _label(linha):
                return linha
        return None

    def _first_exact(texto: str, start: int = 1, end: int | None = None) -> int | None:
        alvo = _texto_normalizado(texto)
        fim = _limite_final(end)
        for linha in range(max(1, start), fim + 1):
            if _label(linha) == alvo:
                return linha
        return None

    row_sources_uses = _first_contains("sources & uses")
    top_end = row_sources_uses - 1 if row_sources_uses else max_row

    row_uses_header = _first_exact("uses", start=(row_sources_uses or 1))
    if row_uses_header is None:
        row_uses_header = _first_contains("uses", start=(row_sources_uses or 1))

    row_total_uses = _first_exact("total", start=(row_uses_header or 1))
    row_exit_values = _first_contains("calculated exit values", start=(row_total_uses or 1))
    row_waterfall = _first_contains("equity waterfall", start=(row_exit_values or row_total_uses or 1))

    uses_start = row_uses_header + 1 if row_uses_header else 1
    uses_end = row_total_uses - 1 if row_total_uses else max_row

    params_start = row_total_uses + 1 if row_total_uses else 1
    if row_exit_values:
        params_end = row_exit_values - 1
    elif row_waterfall:
        params_end = row_waterfall - 1
    else:
        params_end = max_row

    end_year_row = _first_contains("end year", end=top_end)
    down_payment_row = _first_contains("down payment", end=top_end)
    due_pct_row = _first_contains("due diligence costs (%)", end=top_end)
    loan_orig_pct_row = _first_contains("loan origination costs (%)", end=top_end)

    purchase_top_row = _first_contains("purchase price", end=top_end)
    purchase_uses_row = _first_contains("purchase price", start=uses_start, end=uses_end)
    closing_uses_row = _first_contains("closing costs", start=uses_start, end=uses_end)
    immediate_uses_row = _first_contains("immediate repairs", start=uses_start, end=uses_end)
    due_uses_row = _first_contains("due diligence costs", start=uses_start, end=uses_end)
    loan_uses_row = _first_contains("loan origination costs", start=uses_start, end=uses_end)
    acq_fee_uses_row = _first_contains("acquisition fee", start=uses_start, end=uses_end)

    closing_param_row = _first_contains("closing costs", start=params_start, end=params_end)
    immediate_param_row = _first_contains("immediate repairs", start=params_start, end=params_end)
    equity_row = _first_contains("equity contribution", start=params_start, end=params_end)
    cap_partner_row = _first_contains("capital partner contribution", start=params_start, end=params_end)
    manager_row = _first_contains("manager contribution", start=params_start, end=params_end)
    cap_partner_share_row = _first_contains("capital partner share", start=params_start, end=params_end)
    manager_share_row = _first_contains("manager share", start=params_start, end=params_end)
    acq_fee_pct_row = _first_contains("acquisition fee (%", start=params_start, end=params_end)

    seller_amount_row = _first_contains("seller finance amount", end=top_end)
    seller_balloon_row = _first_contains("seller finance balloon", end=top_end)
    acquisition_bank_loan_row = _first_contains("acquisition - bank loan", end=top_end)
    bank_amount_row = None
    if acquisition_bank_loan_row and acquisition_bank_loan_row + 1 <= max_row:
        if "bank loan amount" in _label(acquisition_bank_loan_row + 1):
            bank_amount_row = acquisition_bank_loan_row + 1
    if bank_amount_row is None:
        bank_amount_row = _first_contains("bank loan amount", end=top_end)
    bank_balloon_row = _first_contains("bank loan balloon", end=top_end)

    if seller_amount_row and row_total_uses:
        ws_inputs.cell(row=seller_amount_row, column=3, value=f"=D{row_total_uses}")
    if seller_balloon_row and end_year_row:
        ws_inputs.cell(row=seller_balloon_row, column=3, value=f"=C{end_year_row}")

    if row_total_uses and bank_amount_row is not None:
        # Se a linha de Bank Loan Amount vier mesclada, desfaz merge para manter
        # padrao tabular da Inputs (label em B e valor/formula em C).
        merges_para_desfazer = []
        for merged_range in ws_inputs.merged_cells.ranges:
            if (
                merged_range.min_row <= bank_amount_row <= merged_range.max_row
                and merged_range.min_col <= 2
                and merged_range.max_col >= 3
            ):
                merges_para_desfazer.append(str(merged_range))

        for merged_ref in merges_para_desfazer:
            ws_inputs.unmerge_cells(merged_ref)

        ws_inputs.cell(row=bank_amount_row, column=2, value="Bank Loan Amount")
        ws_inputs.cell(row=bank_amount_row, column=3, value=f"=E{row_total_uses}")

    if bank_balloon_row and end_year_row:
        ws_inputs.cell(row=bank_balloon_row, column=3, value=f"=C{end_year_row}")

    if not purchase_top_row:
        return

    if purchase_uses_row:
        ws_inputs.cell(row=purchase_uses_row, column=3, value=f"=C{purchase_top_row}")

    if down_payment_row:
        ws_inputs.cell(row=down_payment_row, column=4, value=f"=C{purchase_top_row}*C{down_payment_row}")

    if due_pct_row:
        ws_inputs.cell(
            row=due_pct_row,
            column=3,
            value=f"=IF(C{purchase_top_row}=0,0,D{due_pct_row}/C{purchase_top_row})",
        )

    if loan_orig_pct_row:
        ws_inputs.cell(row=loan_orig_pct_row, column=4, value=f"=C{purchase_top_row}*C{loan_orig_pct_row}")

    if closing_uses_row and closing_param_row:
        ws_inputs.cell(row=closing_uses_row, column=3, value=f"=C{closing_param_row}")

    if acq_fee_uses_row and acq_fee_pct_row:
        ws_inputs.cell(row=acq_fee_uses_row, column=3, value=f"=C{purchase_top_row}*C{acq_fee_pct_row}")

    if due_uses_row and due_pct_row:
        ws_inputs.cell(row=due_uses_row, column=3, value=f"=D{due_pct_row}")

    if loan_uses_row and loan_orig_pct_row:
        ws_inputs.cell(row=loan_uses_row, column=3, value=f"=D{loan_orig_pct_row}")

    if immediate_uses_row and immediate_param_row:
        ws_inputs.cell(row=immediate_uses_row, column=3, value=f"=C{immediate_param_row}")

    if row_total_uses and purchase_uses_row and immediate_uses_row:
        ws_inputs.cell(row=row_total_uses, column=3, value=f"=SUM(C{purchase_uses_row}:C{immediate_uses_row})")
        ws_inputs.cell(row=row_total_uses, column=4, value=f"=SUM(D{purchase_uses_row}:D{immediate_uses_row})")
        ws_inputs.cell(row=row_total_uses, column=5, value=f"=SUM(E{purchase_uses_row}:E{immediate_uses_row})")

        for linha in range(purchase_uses_row, immediate_uses_row + 1):
            ws_inputs.cell(row=linha, column=6, value=f"=C{linha}-D{linha}-E{linha}")
        ws_inputs.cell(row=row_total_uses, column=6, value=f"=C{row_total_uses}-D{row_total_uses}-E{row_total_uses}")

    if closing_param_row:
        ws_inputs.cell(row=closing_param_row, column=3, value=f"=C{purchase_top_row}*D{closing_param_row}")

    if equity_row and row_total_uses:
        ws_inputs.cell(row=equity_row, column=3, value=f"=F{row_total_uses}")

    if manager_row and equity_row and cap_partner_row:
        ws_inputs.cell(row=manager_row, column=3, value=f"=C{equity_row}-C{cap_partner_row}")

    if cap_partner_share_row and equity_row and cap_partner_row:
        ws_inputs.cell(row=cap_partner_share_row, column=3, value=f"=IF(C{equity_row}=0,0,C{cap_partner_row}/C{equity_row})")

    if manager_share_row and equity_row and manager_row:
        ws_inputs.cell(row=manager_share_row, column=3, value=f"=IF(C{equity_row}=0,0,C{manager_row}/C{equity_row})")

    refi_cap_rate_row = _first_contains("refi cap rate", end=top_end)
    refi_ltv_row = _first_contains("refi ltv", end=top_end)
    refi_closing_cost_row = _first_contains("refi closing cost", end=top_end)
    refi_year_row = _first_exact("refi year", end=top_end)
    if refi_year_row is None:
        refi_year_row = _first_contains("refi year", end=top_end)
    refi_year_noi_row = _first_contains("refi year noi", end=top_end)
    refi_property_value_row = _first_contains("refi property value", end=top_end)
    refi_closing_cost_amt_row = _first_contains("refi closing costs ($)", end=top_end)
    refi_loan_amount_row = _first_contains("refi loan amount", end=top_end)

    monthly_noi_row = None
    if ws_monthly is not None:
        try:
            monthly_noi_row = _obter_layout_monthly_cf(ws_monthly)["row_noi"]
        except Exception:
            monthly_noi_row = _encontrar_linha_por_texto(ws_monthly, "net operating income", 1)

    monthly_noi_range = (
        f"'Monthly CF'!$C${monthly_noi_row}:$DR${monthly_noi_row}"
        if monthly_noi_row
        else None
    )

    def _formula_noi_por_ano(linha_ano: int) -> str:
        if monthly_noi_range:
            return (
                "=ARRAY_CONSTRAIN(ARRAYFORMULA("
                f"SUMPRODUCT(('Monthly CF'!$C$5:$DR$5=C{linha_ano})*{monthly_noi_range})"
                "),1,1)"
            )

        monthly_noi_series = "INDEX('Monthly CF'!$C:$DR,MATCH(\"Net Operating Income\",'Monthly CF'!$B:$B,0),0)"
        return f"=SUMPRODUCT(('Monthly CF'!$C$5:$DR$5=C{linha_ano})*{monthly_noi_series})"

    if refi_year_noi_row and refi_year_row:
        ws_inputs.cell(
            row=refi_year_noi_row,
            column=3,
            value=_formula_noi_por_ano(refi_year_row),
        )

    if refi_property_value_row and refi_year_noi_row and refi_cap_rate_row:
        ws_inputs.cell(row=refi_property_value_row, column=3, value=f"=C{refi_year_noi_row}/C{refi_cap_rate_row}")

    if refi_closing_cost_amt_row and refi_property_value_row and refi_closing_cost_row:
        ws_inputs.cell(row=refi_closing_cost_amt_row, column=3, value=f"=C{refi_property_value_row}*C{refi_closing_cost_row}")

    if refi_loan_amount_row and refi_property_value_row and refi_ltv_row:
        ws_inputs.cell(row=refi_loan_amount_row, column=3, value=f"=C{refi_property_value_row}*C{refi_ltv_row}")

    exit_cap_rate_row = _first_contains("exit cap rate", start=params_start)
    selling_cost_pct_row = _first_contains("selling cost (%", start=params_start)
    sale_year_noi_row = _first_contains("sale year noi", start=(row_exit_values or params_start))
    selling_price_row = _first_contains("selling price", start=(row_exit_values or params_start))
    selling_cost_amt_row = _first_contains("selling costs ($)", start=(row_exit_values or params_start))
    net_sale_row = _first_contains("net sale proceeds", start=(row_exit_values or params_start))

    if sale_year_noi_row and end_year_row:
        ws_inputs.cell(
            row=sale_year_noi_row,
            column=3,
            value=_formula_noi_por_ano(end_year_row),
        )

    if selling_price_row and sale_year_noi_row and exit_cap_rate_row:
        ws_inputs.cell(row=selling_price_row, column=3, value=f"=C{sale_year_noi_row}/C{exit_cap_rate_row}")

    if selling_cost_amt_row and selling_price_row and selling_cost_pct_row:
        ws_inputs.cell(row=selling_cost_amt_row, column=3, value=f"=C{selling_price_row}*C{selling_cost_pct_row}")

    if net_sale_row and selling_price_row and selling_cost_amt_row:
        ws_inputs.cell(row=net_sale_row, column=3, value=f"=C{selling_price_row}-C{selling_cost_amt_row}")

    gp_catchup_share_row = _first_contains("gp catch-up share", start=(row_waterfall or 1))
    tier_return_capital_row = _first_contains("1. return of capital", start=(row_waterfall or 1))
    tier_pref_return_row = _first_contains("2. preferred return", start=(row_waterfall or 1))
    tier_gp_catchup_row = _first_contains("3. gp catch-up", start=(row_waterfall or 1))
    tier_residual_below_row = _first_contains("4. residual", start=(row_waterfall or 1))
    tier_residual_above_row = _first_contains("5. residual", start=(row_waterfall or 1))

    if tier_return_capital_row:
        ws_inputs.cell(row=tier_return_capital_row, column=4, value=f"=1-C{tier_return_capital_row}")

    if tier_pref_return_row:
        ws_inputs.cell(row=tier_pref_return_row, column=4, value=f"=1-C{tier_pref_return_row}")

    if tier_gp_catchup_row and gp_catchup_share_row:
        ws_inputs.cell(row=tier_gp_catchup_row, column=3, value=f"=1-C{gp_catchup_share_row}")
        ws_inputs.cell(row=tier_gp_catchup_row, column=4, value=f"=C{gp_catchup_share_row}")

    if tier_residual_below_row:
        ws_inputs.cell(row=tier_residual_below_row, column=4, value=f"=1-C{tier_residual_below_row}")

    if tier_residual_above_row:
        ws_inputs.cell(row=tier_residual_above_row, column=4, value=f"=1-C{tier_residual_above_row}")


def aplicar_formulas_apos_inputs(wb, ws_inputs) -> None:
    """API principal do motor: aplica todas as formulas apos preencher Inputs.

Onde atua:
- Inputs e IncomeExpenses (recalculo base).
- Monthly CF (timeline, labels e formulas completas).
- Quarterly CF e Annual CF (expansao + formulas de resumo).
- Summary e Equity Waterfall (links e formulas derivadas).

Por que faz:
- Centraliza num unico passo todo o pos-processamento necessario para o arquivo final.
    """
    purchase_date = _normalizar_purchase_date(ws_inputs["C14"].value)
    end_year = _normalizar_end_year(ws_inputs["C15"].value)
    if ABA_INCOME_EXPENSES not in wb.sheetnames:
        raise ValueError(f"A aba '{ABA_INCOME_EXPENSES}' nao existe no workbook.")
    ws_ie = wb[ABA_INCOME_EXPENSES]

    recalcular_formulas_proforma_inputs(ws_inputs, ws_ie)
    ws_monthly = wb["Monthly CF"] if "Monthly CF" in wb.sheetnames else None

    if ws_monthly is not None:
        nomes_other_income = _extrair_nomes_other_income(ws_ie)
        nomes_expenses = _extrair_nomes_expenses(ws_ie)

        minimo_income_slots = max(BASE_OTHER_INCOME_SLOTS, len(nomes_other_income))
        minimo_expense_slots = max(BASE_EXPENSE_SLOTS, len(nomes_expenses))

        _expandir_linhas_monthly_cf(ws_monthly, minimo_income_slots, minimo_expense_slots)

        for nome_aba in ("Quarterly CF", "Annual CF"):
            if nome_aba in wb.sheetnames:
                _expandir_linhas_resumo_cf(wb[nome_aba], minimo_income_slots, minimo_expense_slots)

    _recalcular_formulas_estrutura_inputs(ws_inputs, ws_monthly)
    _atualizar_formula_summary_exit_cap_rate(wb, ws_inputs)

    if ws_monthly is None:
        return

    total_meses = end_year * 12

    for indice_mes, coluna in enumerate(range(3, ws_monthly.max_column + 1), start=1):
        if purchase_date and indice_mes <= total_meses:
            ws_monthly.cell(row=3, column=coluna, value=indice_mes)
            ws_monthly.cell(row=4, column=coluna, value=((indice_mes - 1) // 3) + 1)
            ws_monthly.cell(row=5, column=coluna, value=((indice_mes - 1) // 12) + 1)

            mes_base = (purchase_date.month - 1) + (indice_mes - 1)
            ano = purchase_date.year + (mes_base // 12)
            mes = (mes_base % 12) + 1
            ws_monthly.cell(row=6, column=coluna, value=date(ano, mes, 1))
        else:
            ws_monthly.cell(row=3, column=coluna, value=None)
            ws_monthly.cell(row=4, column=coluna, value=None)
            ws_monthly.cell(row=5, column=coluna, value=None)
            ws_monthly.cell(row=6, column=coluna, value=None)

    layout_monthly = _obter_layout_monthly_cf(ws_monthly)
    _atualizar_nomes_other_income_monthly_cf(
        ws_monthly,
        nomes_other_income,
        linha_inicial=layout_monthly["income_start"],
        linha_final=layout_monthly["income_end"],
    )
    _atualizar_nomes_expenses_monthly_cf(
        ws_monthly,
        nomes_expenses,
        linha_inicial=layout_monthly["expense_start"],
        linha_final=layout_monthly["expense_end"],
    )

    _aplicar_formulas_monthly_cf_dinamicas(ws_monthly, ws_inputs, ws_ie)
    _atualizar_labels_resumo_cf(wb)
    _atualizar_formulas_equity_waterfall(wb)


def reaplicar_formulas_do_template(
    arquivo_destino,
    arquivo_template: Path | str = ARQUIVO_TEMPLATE,
    nome_aba: str = ABA_PADRAO,
):
    """Copia formulas da aba de template para arquivo destino e recalcula dinamicos.

Onde atua:
- Copia formulas da aba definida (padrao: Monthly CF) do InputTemplate.xlsx.
- Em seguida roda o pipeline dinamico quando houver aba Inputs no destino.

Por que faz:
- Permite recuperar formulas perdidas/alteradas em arquivos da pasta Output.
    """
    destino_path = Path(arquivo_destino)
    template_path = Path(arquivo_template)

    if not template_path.exists():
        raise FileNotFoundError(f"Template nao encontrado: {template_path}")
    if not destino_path.exists():
        raise FileNotFoundError(f"Arquivo destino nao encontrado: {destino_path}")

    wb_template = load_workbook(template_path, data_only=False)
    wb_destino = load_workbook(destino_path, data_only=False)

    if nome_aba not in wb_template.sheetnames:
        raise ValueError(f"A aba '{nome_aba}' nao existe no template.")
    if nome_aba not in wb_destino.sheetnames:
        raise ValueError(f"A aba '{nome_aba}' nao existe no arquivo destino.")

    ws_template = wb_template[nome_aba]
    ws_destino = wb_destino[nome_aba]
    ws_inputs_destino = wb_destino["Inputs"] if "Inputs" in wb_destino.sheetnames else None

    total_formulas = 0
    for row in ws_template.iter_rows(
        min_row=1,
        max_row=ws_template.max_row,
        min_col=1,
        max_col=ws_template.max_column,
    ):
        for cell in row:
            valor = cell.value
            if isinstance(valor, str) and valor.startswith("="):
                _set_cell_value_respeitando_merge(ws_destino, cell.coordinate, valor)
                total_formulas += 1

    if ws_inputs_destino is not None and nome_aba == ABA_PADRAO:
        aplicar_formulas_apos_inputs(wb_destino, ws_inputs_destino)

    wb_destino.save(destino_path)
    wb_template.close()
    wb_destino.close()

    return total_formulas


def _listar_xlsx(path_output) -> set[Path]:
    """Lista arquivos .xlsx validos na pasta Output, ignorando temporarios do Excel."""
    path_output = Path(path_output)
    if not path_output.exists():
        path_output.mkdir(parents=True, exist_ok=True)

    return {
        arquivo.resolve()
        for arquivo in path_output.glob("*.xlsx")
        if arquivo.is_file() and not arquivo.name.startswith("~$")
    }


def processar_output_uma_vez(
    path_output: Path | str = PASTA_OUTPUT,
    arquivo_template: Path | str = ARQUIVO_TEMPLATE,
    nome_aba: str = ABA_PADRAO,
) -> None:
    """Executa a reaplicacao de formulas em lote na pasta Output.

Onde atua:
- Todos os .xlsx da pasta Output.

Por que faz:
- Facilita manutencao em lote sem abrir arquivo por arquivo manualmente.
    """
    path_output = Path(path_output)
    arquivo_template = Path(arquivo_template)

    print(f"Verificando pasta: {path_output.resolve()}")
    print(f"Template de formulas: {arquivo_template.resolve()}\n")

    arquivos = sorted(_listar_xlsx(path_output))
    if not arquivos:
        print("Nenhum arquivo .xlsx encontrado na pasta Output.")
        return

    for arquivo in arquivos:
        print(f"Processando: {arquivo.name}")
        try:
            qtd = reaplicar_formulas_do_template(
                arquivo_destino=arquivo,
                arquivo_template=arquivo_template,
                nome_aba=nome_aba,
            )
            print(f"OK | {arquivo.name} | formulas aplicadas: {qtd}")
        except Exception as erro:
            print(f"ERRO | {arquivo.name} | {erro}")


def main() -> None:
    """Entrypoint CLI para o modo de manutencao de formulas na Output."""
    parser = argparse.ArgumentParser(
        description="Verifica a pasta Output e reaplica formulas nos arquivos Excel."
    )
    parser.add_argument(
        "--output",
        default=str(PASTA_OUTPUT),
        help="Pasta monitorada (padrao: Output)",
    )
    parser.add_argument(
        "--template",
        default=str(ARQUIVO_TEMPLATE),
        help="Arquivo template com formulas (padrao: InputTemplate.xlsx)",
    )
    parser.add_argument(
        "--sheet",
        default=ABA_PADRAO,
        help="Nome da aba para copiar formulas (padrao: Monthly CF)",
    )

    args = parser.parse_args()
    processar_output_uma_vez(
        path_output=args.output,
        arquivo_template=args.template,
        nome_aba=args.sheet,
    )


if __name__ == "__main__":
    main()
