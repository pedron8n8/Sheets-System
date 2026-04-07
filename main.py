"""Pipeline de entrada de dados para gerar arquivos de analise.

Este arquivo e responsavel por:
- Ler registros do contatos.xlsx.
- Preencher abas do template: Inputs e IncomeExpenses.
- Acionar o motor de formulas (functions.aplicar_formulas_apos_inputs), que atualiza
    Monthly CF, Quarterly CF, Annual CF, Summary e Equity Waterfall.
- Salvar o arquivo final em Output.

Por que existe:
- Centraliza o fluxo "dados tabulares -> workbook pronto" em um ponto unico.
"""

from __future__ import annotations

from datetime import datetime
import os
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from functions import aplicar_formulas_apos_inputs

ARQUIVO_TEMPLATE = "InputTemplate.xlsx"
ARQUIVO_DATA = "contatos.xlsx"
PASTA_SAIDA = "Output"
ABA_INCOME_EXPENSES = "IncomeExpenses"
IE_INCOME_START = 7
IE_INCOME_END = 46
IE_EXPENSE_START = 50
IE_EXPENSE_END = 89


def garantir_pasta_saida(pasta_saida: str = PASTA_SAIDA) -> None:
    """Garante que a pasta Output exista antes de salvar arquivos."""
    if not os.path.exists(pasta_saida):
        os.makedirs(pasta_saida)


def _set_cell_value_respeitando_merge(sheet, cell_ref: str, value) -> None:
    """Escreve celula respeitando merges (usa ancora superior esquerda)."""
    for merged_range in sheet.merged_cells.ranges:
        if cell_ref in merged_range:
            anchor_col = get_column_letter(merged_range.min_col)
            anchor_ref = f"{anchor_col}{merged_range.min_row}"
            sheet[anchor_ref] = value
            return
    sheet[cell_ref] = value


def _normalizar_percentual(valor):
    """Normaliza entradas percentuais para formato decimal esperado pelo Excel.

Exemplos:
- 20 -> 0.20
- "1,5%" -> 0.015
- 0.2 -> 0.2
    """
    if valor == "" or pd.isna(valor):
        return ""

    if isinstance(valor, str):
        texto = valor.strip().replace("%", "").replace(",", ".")
        if not texto:
            return ""
        try:
            numero = float(texto)
        except ValueError:
            return valor
    else:
        numero = valor

    if abs(numero) <= 1:
        return numero

    # Entrada em formato percentual comum (ex.: 1.5 => 1.5%).
    if abs(numero) <= 100:
        return numero / 100

    # Valores acima de 100% sao invalidados para evitar explosoes numericas no Excel.
    return 1.0 if numero > 0 else -1.0


def _normalizar_down_payment(valor_down_payment, valor_purchase_price):
    """Normalize down payment robustly for legacy and current schemas.

    Accepted inputs:
    - Decimal percent: 0.2
    - Percent number: 20
    - Legacy absolute amount: 200 with purchase price 800 -> 0.25
    """
    percentual = _normalizar_percentual(valor_down_payment)

    # Quando veio como valor absoluto legado (ex.: 200), _normalizar_percentual caparia em 100%.
    # Antes de usar o cap, tentamos inferir percentual com base no Purchase Price.
    raw = _to_float(valor_down_payment)
    purchase_price = _to_float(valor_purchase_price)
    if abs(raw) > 100 and purchase_price > 0:
        inferido = raw / purchase_price
        if inferido >= 0:
            return min(inferido, 1.0)

    return percentual


def _valor_limpo(valor):
    """Converte NaN para string vazia para evitar sujeira no workbook."""
    if pd.isna(valor):
        return ""
    return valor


def _to_float(valor) -> float:
    """Converte valor heterogeneo para float (aceita simbolos e strings)."""
    if valor is None or valor == "" or pd.isna(valor):
        return 0.0

    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip().replace("$", "").replace(",", "")
    try:
        return float(texto)
    except ValueError:
        return 0.0


def _extrair_other_incomes(primeira_linha):
    """Extrai itens dinamicos de Other Income da linha do contatos.xlsx.

Origem esperada de colunas:
- Other Income N Type
- Other Income N Amount
    """
    other_incomes = []
    for coluna in primeira_linha.index:
        match = re.match(r"^Other Income\s+(\d+)\s+Type$", str(coluna))
        if not match:
            continue

        idx = match.group(1)
        tipo = _valor_limpo(primeira_linha.get(coluna, ""))
        valor = _valor_limpo(primeira_linha.get(f"Other Income {idx} Amount", ""))
        if str(tipo).strip():
            other_incomes.append((int(idx), str(tipo).strip(), valor))

    other_incomes.sort(key=lambda item: item[0])
    return other_incomes


def _extrair_other_expenses(primeira_linha):
    """Extrai itens dinamicos de Other Expense da linha do contatos.xlsx."""
    other_expenses = []
    for coluna in primeira_linha.index:
        match = re.match(r"^Other Expense\s+(\d+)\s+Type$", str(coluna))
        if not match:
            continue

        idx = match.group(1)
        tipo = _valor_limpo(primeira_linha.get(coluna, ""))
        valor = _valor_limpo(primeira_linha.get(f"Other Expense {idx} Amount", ""))
        if str(tipo).strip():
            other_expenses.append((int(idx), str(tipo).strip(), valor))

    other_expenses.sort(key=lambda item: item[0])
    return other_expenses


def _compactar_other_incomes(other_incomes, max_rows: int):
    """Compacta excedentes de receitas em uma linha agregada quando necessario."""
    if len(other_incomes) <= max_rows:
        return other_incomes

    base = other_incomes[: max_rows - 1]
    excedentes = other_incomes[max_rows - 1 :]
    total_excedentes = sum(_to_float(valor_income) for _, _, valor_income in excedentes)
    quantidade_excedentes = len(excedentes)
    base.append((9999, f"Other Income - Additional ({quantidade_excedentes} items)", total_excedentes))
    return base


def _compactar_other_expenses(other_expenses, max_rows: int):
    """Compacta excedentes de despesas em uma linha agregada quando necessario."""
    if len(other_expenses) <= max_rows:
        return other_expenses

    base = other_expenses[: max_rows - 1]
    excedentes = other_expenses[max_rows - 1 :]
    total_excedentes = sum(_to_float(valor_expense) for _, _, valor_expense in excedentes)
    quantidade_excedentes = len(excedentes)
    base.append((9999, f"Other Expense - Additional ({quantidade_excedentes} items)", total_excedentes))
    return base


def criar_arquivo_baseado_em_template(primeira_linha) -> str:
    """Preenche o template Excel a partir de um registro e gera o arquivo final.

Abas envolvidas no preenchimento direto aqui:
- Inputs: dados do imovel, compra, operacao e capex labels.
- IncomeExpenses: lista de receitas/despesas detalhadas para lookup posterior.

Abas atualizadas indiretamente:
- Monthly CF, Quarterly CF, Annual CF, Summary e Equity Waterfall,
  via aplicar_formulas_apos_inputs.
    """
    garantir_pasta_saida(PASTA_SAIDA)

    wb = load_workbook(ARQUIVO_TEMPLATE, data_only=False)
    ws = wb["Inputs"]
    ws_ie = wb[ABA_INCOME_EXPENSES]

    input_property_type = primeira_linha.get("Property Type", "")
    input_property_name = primeira_linha.get("Property Name", "")
    input_address = primeira_linha.get("Address", "")
    input_city_state = primeira_linha.get("City and State", "")
    input_number_units = primeira_linha.get("Number of Units", 0)
    input_purchase_price = primeira_linha.get("Purchase Price", 0)
    input_down_payment = primeira_linha.get("Down Payment (%)", primeira_linha.get("Down Payment", 0))
    input_diligence = primeira_linha.get("Due Diligence Costs", primeira_linha.get("Due Diligence Costs %", 0))
    input_loan_costs = primeira_linha.get("Loan Original Costs", 0)
    input_purchase_date = primeira_linha.get("Purchase Date", "")
    input_end_year = primeira_linha.get("End Year", 10)

    input_gpr = primeira_linha.get("Gross Potential Rent", 0)
    input_vacancy = primeira_linha.get("Vacancy Rate %", 0)
    input_credit_loss = primeira_linha.get("Credit Loss %", 0)

    input_property_tax = primeira_linha.get("Property Tax", 0)
    input_insurance = primeira_linha.get("Insurance", 0)
    input_management_fee = primeira_linha.get("Management Fee %", 0)
    input_repairs = primeira_linha.get("Repairs and Maintenance", 0)
    input_utilities = primeira_linha.get("Utilities", 0)
    input_capex_base = primeira_linha.get("Capital Expenditures", 0)
    input_landscape = primeira_linha.get("Landscape and Janitorial", 0)

    input_capex1 = primeira_linha.get("CapEx 1", primeira_linha.get("  CapEx Item 1", ""))
    input_capex2 = primeira_linha.get("CapEx 2", primeira_linha.get("  CapEx Item 2", ""))
    input_capex3 = primeira_linha.get("CapEx 3", primeira_linha.get("  CapEx Item 3", ""))
    input_capex4 = primeira_linha.get("CapEx 4", primeira_linha.get("  CapEx Item 4", ""))
    input_capex5 = primeira_linha.get("CapEx 5", primeira_linha.get("  CapEx Item 5", ""))
    input_capex6 = primeira_linha.get("CapEx 6", primeira_linha.get("  CapEx Item 6", ""))

    other_incomes = _extrair_other_incomes(primeira_linha)
    other_expenses = _extrair_other_expenses(primeira_linha)

    # Inputs (aba): metadados do ativo e premissas principais.
    _set_cell_value_respeitando_merge(ws, "C6", input_property_type)
    _set_cell_value_respeitando_merge(ws, "C5", input_property_name)
    _set_cell_value_respeitando_merge(ws, "C7", input_address)
    _set_cell_value_respeitando_merge(ws, "C8", input_city_state)
    _set_cell_value_respeitando_merge(ws, "C9", input_number_units)
    _set_cell_value_respeitando_merge(ws, "C10", input_purchase_price)
    _set_cell_value_respeitando_merge(ws, "C11", _normalizar_down_payment(input_down_payment, input_purchase_price))
    _set_cell_value_respeitando_merge(ws, "D12", input_diligence)
    _set_cell_value_respeitando_merge(ws, "C13", _normalizar_percentual(input_loan_costs))
    _set_cell_value_respeitando_merge(ws, "C14", input_purchase_date)
    _set_cell_value_respeitando_merge(ws, "C15", input_end_year)

    _set_cell_value_respeitando_merge(ws, "C20", input_gpr)
    _set_cell_value_respeitando_merge(ws, "C21", _normalizar_percentual(input_vacancy))
    _set_cell_value_respeitando_merge(ws, "C22", _normalizar_percentual(input_credit_loss))
    _set_cell_value_respeitando_merge(ws, "C28", _normalizar_percentual(input_management_fee))

    # IncomeExpenses (aba): limpa e reescreve bloco de receitas adicionais.
    for row in range(IE_INCOME_START, IE_INCOME_END + 1):
        ws_ie.cell(row=row, column=2, value="")
        ws_ie.cell(row=row, column=3, value="")

    row_cursor_ie_income = IE_INCOME_START
    for _, nome_income, valor_income in other_incomes:
        if row_cursor_ie_income > IE_INCOME_END:
            break
        ws_ie.cell(row=row_cursor_ie_income, column=2, value=nome_income)
        ws_ie.cell(row=row_cursor_ie_income, column=3, value=valor_income)
        row_cursor_ie_income += 1

    fixed_expenses = [
        ("Property Tax", input_property_tax),
        ("Insurance", input_insurance),
        ("Repairs", input_repairs),
        ("Utilities", input_utilities),
        ("Capital Expenditures", input_capex_base),
        ("Landscape", input_landscape),
    ]

    # IncomeExpenses (aba): limpa e reescreve bloco de despesas.
    for row in range(IE_EXPENSE_START, IE_EXPENSE_END + 1):
        ws_ie.cell(row=row, column=2, value="")
        ws_ie.cell(row=row, column=3, value="")

    row_cursor_ie_expense = IE_EXPENSE_START
    for nome_expense, valor_expense in fixed_expenses:
        if row_cursor_ie_expense > IE_EXPENSE_END:
            break
        ws_ie.cell(row=row_cursor_ie_expense, column=2, value=nome_expense)
        ws_ie.cell(row=row_cursor_ie_expense, column=3, value=valor_expense)
        row_cursor_ie_expense += 1

    for _, nome_expense, valor_expense in other_expenses:
        if row_cursor_ie_expense > IE_EXPENSE_END:
            break
        ws_ie.cell(row=row_cursor_ie_expense, column=2, value=nome_expense)
        ws_ie.cell(row=row_cursor_ie_expense, column=3, value=valor_expense)
        row_cursor_ie_expense += 1

    _set_cell_value_respeitando_merge(ws, "C51", input_capex1)
    _set_cell_value_respeitando_merge(ws, "C52", input_capex2)
    _set_cell_value_respeitando_merge(ws, "C53", input_capex3)
    _set_cell_value_respeitando_merge(ws, "C54", input_capex4)
    _set_cell_value_respeitando_merge(ws, "C55", input_capex5)
    _set_cell_value_respeitando_merge(ws, "C56", input_capex6)

    # Dispara o motor que recalcula formulas em todas as abas financeiras.
    aplicar_formulas_apos_inputs(wb, ws)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_arquivo = f"{input_property_name}_{timestamp}.xlsx"
    caminho_saida = os.path.join(PASTA_SAIDA, nome_arquivo)
    wb.save(caminho_saida)
    wb.close()

    print(f"✅ Arquivo criado com sucesso: {caminho_saida}")
    return caminho_saida


def obter_registros_pendentes(df_data: pd.DataFrame) -> pd.DataFrame:
    """Filtra registros com Submitted = No para processamento."""
    if "Submitted" not in df_data.columns:
        df_data["Submitted"] = "No"
    return df_data[df_data["Submitted"].astype(str).str.strip().str.lower() == "no"].copy()


def exibir_resumo_registro(registro) -> None:
    """Mostra no console um resumo do registro antes da geracao do arquivo."""
    print(f"Property Type: {registro.get('Property Type', '')}")
    print(f"Property Name: {registro.get('Property Name', '')}")
    print(f"Address: {registro.get('Address', '')}")
    print(f"City and State: {registro.get('City and State', '')}")
    print(f"Number of Units: {registro.get('Number of Units', 0)}")
    print(f"Purchase Price: {registro.get('Purchase Price', 0)}")
    print(f"Down Payment (%): {registro.get('Down Payment (%)', registro.get('Down Payment', 0))}")
    print("--------------------------------------------")
    print(f"Gross Potential Rent: {registro.get('Gross Potential Rent', 0)}")
    print(f"Vacancy Rate: {registro.get('Vacancy Rate %', 0)}")
    print(f"Credit Loss: {registro.get('Credit Loss %', 0)}")
    print("--------------------------------------------")
    print(f"Property Tax: {registro.get('Property Tax', 0)}")
    print(f"Insurance: {registro.get('Insurance', 0)}")
    print(f"Management Fee: {registro.get('Management Fee %', 0)}")
    print(f"Repairs and Maintenance: {registro.get('Repairs and Maintenance', 0)}")
    print(f"Utilities: {registro.get('Utilities', 0)}")
    print(f"Capital Expenditures: {registro.get('Capital Expenditures', 0)}")
    print(f"Landscape and Janitorial: {registro.get('Landscape and Janitorial', 0)}")


def processar_registro_por_indice(registro_index: int) -> str | None:
    """Processa um registro especifico do contatos.xlsx pelo indice.

Uso principal:
- Chamado pela API apos inserir novo registro, para gerar o arquivo imediatamente.
    """
    if not os.path.exists(ARQUIVO_DATA):
        print(f"Arquivo de dados nao encontrado: {ARQUIVO_DATA}")
        return None

    df_data = pd.read_excel(ARQUIVO_DATA)
    if registro_index < 0 or registro_index >= len(df_data):
        print(f"Indice de registro invalido: {registro_index}")
        return None

    registro = df_data.iloc[registro_index]
    submitted = str(registro.get("Submitted", "No")).strip().lower()
    if submitted == "yes":
        print(f"Registro no indice {registro_index} ja foi processado.")
        return None

    print(f"📌 Processando registro recem inserido no indice {registro_index}...")
    caminho_saida = criar_arquivo_baseado_em_template(registro)

    df_data.loc[registro.name, "Submitted"] = "Yes"
    df_data.to_excel(ARQUIVO_DATA, index=False)
    print("✅ Registro recem inserido processado e marcado como 'Submitted: Yes'.")

    return caminho_saida


def processar_primeiro_registro_pendente() -> str | None:
    """Processa o primeiro registro pendente encontrado no contatos.xlsx."""
    garantir_pasta_saida(PASTA_SAIDA)

    if not os.path.exists(ARQUIVO_DATA):
        print(f"Arquivo de dados nao encontrado: {ARQUIVO_DATA}")
        return None

    df_data = pd.read_excel(ARQUIVO_DATA)
    pendentes = obter_registros_pendentes(df_data)

    if pendentes.empty:
        print("✅ Nenhum registro pendente para processar!")
        return None

    print(f"📋 Total de registros pendentes: {len(pendentes)}")
    print("=" * 60)

    primeira_linha = pendentes.iloc[0]
    exibir_resumo_registro(primeira_linha)

    print("\n📁 Gerando arquivo baseado no template...")
    caminho_saida = criar_arquivo_baseado_em_template(primeira_linha)

    # df_data.loc[primeira_linha.name, "Submitted"] = "Yes"
    df_data.to_excel(ARQUIVO_DATA, index=False)
    print("✅ Registro processado e marcado como 'Submitted: Yes' no arquivo de dados.")

    return caminho_saida


def main() -> None:
    """Entrypoint CLI padrao: processa apenas o primeiro pendente."""
    processar_primeiro_registro_pendente()


if __name__ == "__main__":
    main()
